import dash
from dash import dash_table, dcc, html, Input, Output, State, ctx
import dash_bootstrap_components as dbc
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import base64
import io
import unidecode
from datetime import datetime
import re
import numpy as np

image_path = 'assets/logo.png'

items = [
    dbc.DropdownMenuItem("GROUPE DE RESOLUTION", id="item-1"),
    dbc.DropdownMenuItem("AGENCE", id="item-2"),
    dbc.DropdownMenuItem("TICKET", id="item-3"),
]

app = dash.Dash(__name__, external_stylesheets=[dbc.themes.CYBORG], suppress_callback_exceptions=True)
server = app.server  # Expose le serveur Flask pour Gunicorn

# Augmenter la limite de taille des fichiers uploadés (défaut ~2MB, ici 50MB)
server.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# ── Palette centralisée
C = {
    "bg":      "#0d1117", "sidebar": "#161b22", "card":   "#1c2128",
    "card2":   "#21262d", "accent":  "#58a6ff", "green":  "#3fb950",
    "red":     "#f78166", "orange":  "#ffa657", "muted":  "#8b949e",
    "text":    "#e6edf3", "border":  "#30363d", "violet": "#6e40c9",
}
PLOTLY_BASE = dict(
    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
    font=dict(color="#e6edf3", family="'Segoe UI', sans-serif"),
    margin=dict(l=12, r=12, t=48, b=12),
    hoverlabel=dict(bgcolor="#21262d", font_color="#e6edf3", bordercolor="#30363d"),
    xaxis=dict(showgrid=False, zeroline=False, color="#8b949e", tickfont=dict(size=11)),
    yaxis=dict(showgrid=True, gridcolor="#30363d", zeroline=False, color="#8b949e", tickfont=dict(size=11)),
)
def apply_layout(fig, **kw):
    cfg = dict(PLOTLY_BASE); cfg.update(kw); fig.update_layout(**cfg); return fig

card_style  = {"borderRadius": "12px", "border": f"1px solid #30363d",
               "backgroundColor": "#1c2128", "padding": "16px", "marginBottom": "16px"}
btn_style   = {"borderRadius": "8px", "fontWeight": "600", "fontSize": "13px",
               "border": "none", "cursor": "pointer", "letterSpacing": "0.3px"}
label_style = {"color": "#8b949e", "fontWeight": "500", "fontSize": "11px", "marginBottom": "5px",
               "display": "block", "textTransform": "uppercase", "letterSpacing": "0.8px"}
title_style = {"color": "#e6edf3", "fontWeight": "700", "marginBottom": "15px"}

def kpi_card(label, value, icon, color, suffix=""):
    return dbc.Col(html.Div([
        html.Div([
            html.Span(icon, style={"fontSize":"20px","marginRight":"8px"}),
            html.Span(label, style={"fontSize":"11px","color":"#8b949e","textTransform":"uppercase",
                                    "letterSpacing":"0.8px","fontWeight":"600"}),
        ], style={"display":"flex","alignItems":"center","marginBottom":"8px"}),
        html.Div(f"{value}{suffix}", style={"fontSize":"26px","fontWeight":"800","color":color,"lineHeight":"1"}),
    ], style={
        "backgroundColor":"#1c2128","border":"1px solid #30363d",
        "borderLeft":f"3px solid {color}","borderRadius":"10px",
        "padding":"14px 18px","marginBottom":"16px",
    }), width=3)

def kpi_row(df):
    total = df["Complaint reference"].nunique() if "Complaint reference" in df.columns else 0
    hd    = (df["DELAI_RECLAMATION"]=="HORS DELAI").sum() if "DELAI_RECLAMATION" in df.columns else 0
    taux  = round(hd/total*100,1) if total else 0
    fond  = (df["NATURE"]=="FONDEE").sum() if "NATURE" in df.columns else 0
    return dbc.Row([
        kpi_card("Total réclamations", f"{total:,}".replace(","," "), "📋", "#58a6ff"),
        kpi_card("Hors délai",         f"{hd:,}".replace(","," "),   "⚠️", "#f78166"),
        kpi_card("Taux hors délai",    taux,                          "📊", "#ffa657", "%"),
        kpi_card("Fondées",            f"{fond:,}".replace(","," "), "✅", "#3fb950"),
    ], className="mb-2")

DURATION_PATTERNS = {
    'days': re.compile(r'(\d+)\s*[dDjJ]'), 'hours': re.compile(r'(\d+)\s*[hH]'),
    'minutes': re.compile(r'(\d+)\s*[mM]'), 'seconds': re.compile(r'(\d+)\s*[sS]')
}

def convert_to_days(duration):
    if pd.isna(duration) or duration == '': return 0
    total_seconds = 0
    for unit, pattern in DURATION_PATTERNS.items():
        match = pattern.search(duration)
        if match:
            value = int(match.group(1))
            if unit == 'days': total_seconds += value * 86400
            elif unit == 'hours': total_seconds += value * 3600
            elif unit == 'minutes': total_seconds += value * 60
            elif unit == 'seconds': total_seconds += value
    return round(total_seconds / 86400, 2)

def categorize_segment(segment):
    if str(segment).startswith('101'): return 'PARTICULIER'
    elif str(segment).startswith('102'): return 'PROFESSIONNEL'
    elif segment == 'INCONNU': return 'INCONNU'
    else: return 'CORPORATE'

def load_data(df):
    df.columns = df.columns.str.replace("\n", " ")
    df["Claim SLA"] = df["Claim SLA"].str.replace('[', "", regex=False).str.replace('REC', "", regex=False).str.replace('-', "", regex=False).str.replace(']', "", regex=False)
    df['Claim SLA'] = df['Claim SLA'].str.split(',').explode().reset_index(drop=True)
    df['Claim SLA'] = df['Claim SLA'].str.strip()
    df['SLA_ETAPE'] = df['Claim SLA'].apply(lambda x: x.split(':')[0].strip() if ':' in x else None)
    df['Value'] = df['Claim SLA'].apply(lambda x: x.split(':')[1].strip() if ':' in x else None)
    df['SEGMENTATION'] = df['Segment'].apply(categorize_segment)
    df["Nature of the claim"] = df["Nature of the claim"].fillna("INCONNU")
    df.loc[df["Nature of the claim"] == 'Fondé avec faute SG', "NATURE"] = 'FONDEE'
    df.loc[df["Nature of the claim"] == 'Fondé sans faute SG', "NATURE"] = 'FONDEE'
    df.loc[df["Nature of the claim"] == 'Non fondée', "NATURE"] = 'NON FONDEE'
    df['NATURE'] = df['NATURE'].fillna('INCONNU')
    df['Date of creation'] = pd.to_datetime(df['Date of creation'].str.split(' ').str[0], format='%d-%m-%Y', errors='coerce')
    df['Resolved date'] = pd.to_datetime(df['Resolved date'].str.split(' ').str[0], format='%d-%m-%Y', errors='coerce')
    df['Annee'] = df['Date of creation'].dt.year
    df['Mois'] = df['Date of creation'].dt.month
    df["GROUPE RESOLUTION"] = df["Resolution group"].str.replace('SGCI', "", regex=False)
    df["AGENCE"] = df["Branch"].str[6:]
    df["Typology"] = df.Typology.str.upper().apply(unidecode.unidecode).str.replace("'", " ", regex=False)
    df["DATE_AUJOURDUI"] = datetime.today()
    time_columns = ["Time Technical Study", "Time Additional information", "Time Treatment",
                    "Time SUPPORT", "Time Treated", "Time To Complete", "Time Initialization",
                    "Time Validate Regularisation", "Time In the process of regularization", "Time Third party return waiting"]
    for col in time_columns:
        if col in df.columns: df[col] = pd.to_numeric(df[col], errors='coerce') / 86400
        else: df[col] = 0
    df["Treatment duration"] = df['Treatment duration (In Days)'].apply(convert_to_days) if 'Treatment duration (In Days)' in df.columns else 0
    df["Duree Traitee"] = df['Value'].apply(convert_to_days)
    df['signe'] = np.where(df["Value"].str[0] == "-", -1, 1)
    df["Duree Traitee"] = pd.to_numeric(df["Duree Traitee"], errors='coerce').fillna(0).astype(int)
    df["signe"] = df["signe"].astype(int)
    df["SLA_JOURS"] = df["Duree Traitee"] * df["signe"]
    df['DATE_RECLAMATION'] = df['Resolved date'].where(df['Resolved date'].notna(), df['DATE_AUJOURDUI']).sub(df['Date of creation']).dt.days
    df['DELAI_RECLAMATION'] = np.where(df['DATE_RECLAMATION'] <= 30, 'PAS HORS DELAI', 'HORS DELAI')
    return df

def parse_contents(contents, filename):
    try:
        content_type, content_string = contents.split(',', 1)
        decoded = base64.b64decode(content_string)
        if filename.endswith('.csv'):
            # Essayer UTF-8, puis latin-1 si ça échoue
            try:
                df = pd.read_csv(io.StringIO(decoded.decode('utf-8')), header=7)
            except UnicodeDecodeError:
                df = pd.read_csv(io.StringIO(decoded.decode('latin-1')), header=7)
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            df = pd.read_excel(io.BytesIO(decoded), engine='openpyxl', header=7)
        else:
            return None, f"Type de fichier non pris en charge : {filename}"
        if df.empty:
            return None, "Le fichier est vide."
        return df, ""
    except Exception as e:
        return None, f"Erreur lors du chargement : {str(e)}"

def create_datatable_from_df(df, page_size=10):
    def style_sla_cells(dataframe):
        styles = []
        for i, row in dataframe.iterrows():
            for col in dataframe.columns:
                if col == dataframe.columns[0]: continue
                value = row[col]
                if pd.isnull(value): continue
                try: val_float = float(value)
                except: continue
                if val_float > 0: styles.append({'if': {'row_index': i, 'column_id': col}, 'color': '#2ecc71', 'fontWeight': 'bold'})
                elif val_float < 0: styles.append({'if': {'row_index': i, 'column_id': col}, 'color': '#e74c3c', 'fontWeight': 'bold'})
        return styles
    return dash_table.DataTable(
        columns=[{"name": i, "id": i} for i in df.columns], data=df.to_dict('records'), page_size=page_size,
        style_table={'overflowX': 'auto'},
        style_cell={'textAlign': 'left', 'minWidth': '120px', 'whiteSpace': 'normal',
                    'color': 'white', 'backgroundColor': '#22272e', 'border': 'none'},
        style_header={'backgroundColor': '#2b303a', 'fontWeight': 'bold', 'color': 'white', 'border': 'none'},
        style_data_conditional=style_sla_cells(df), fill_width=True)

def add_line_breaks(text, max_chars=10):
    words = text.split(); line = ""; new_text = ""
    for word in words:
        if len(line) + len(word) + 1 <= max_chars: line += (word + " ")
        else: new_text += line.rstrip() + "<br>"; line = word + " "
    new_text += line.rstrip()
    return new_text

def generate_view_1(data):
    graphs = [dbc.Col(kpi_row(data), width=12)]
    # Donut nature
    nat = data.groupby("NATURE")["Complaint reference"].count().reset_index(name='nombre')
    fig1 = px.pie(nat, values='nombre', names='NATURE', title="Répartition par nature",
                  hole=0.55, color_discrete_sequence=["#58a6ff","#3fb950","#f78166","#ffa657"], template='plotly_dark')
    fig1.update_traces(textinfo='percent+label', hovertemplate="<b>%{label}</b><br>%{value}<extra></extra>")
    apply_layout(fig1)
    fig1.update_layout(legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5))
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig1, config={"displayModeBar":False}), style=card_style), width=6))
    # Line évolution avec aire
    ev = data.groupby("Mois")["Complaint reference"].nunique().reset_index(name='nombre')
    ev['Mois'] = ev['Mois'].apply(lambda m: datetime(2000, m, 1).strftime('%b'))
    fig2 = go.Figure(go.Scatter(
        x=ev['Mois'], y=ev['nombre'], mode='lines+markers+text',
        text=ev['nombre'], textposition='top center',
        line=dict(color="#58a6ff", width=2.5), marker=dict(size=8, color="#58a6ff"),
        fill='tozeroy', fillcolor="rgba(88,166,255,0.08)",
        hovertemplate="<b>%{x}</b><br>%{y} réclamations<extra></extra>",
    ))
    fig2.update_layout(title="Évolution mensuelle des réclamations")
    apply_layout(fig2)
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig2, config={"displayModeBar":False}), style=card_style), width=6))
    # Bar agences
    ag = data.groupby("AGENCE")["Complaint reference"].count().reset_index(name='nombre').sort_values('nombre').tail(10)
    fig3 = px.bar(ag, y="AGENCE", x='nombre', text='nombre', orientation='h',
                  title='Top 10 agences', color='nombre',
                  color_continuous_scale=px.colors.sequential.Blues, template='plotly_dark')
    fig3.update_traces(textfont_size=12, marker_line_width=0,
                       hovertemplate="<b>%{y}</b><br>%{x} réclamations<extra></extra>")
    fig3.update_layout(coloraxis_showscale=False)
    apply_layout(fig3, xaxis=dict(visible=False))
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig3, config={"displayModeBar":False}), style=card_style), width=12))
    return graphs

def generate_view_2(data):
    graphs = [dbc.Col(kpi_row(data), width=12)]
    gr = data.groupby("GROUPE RESOLUTION")["Complaint reference"].count().reset_index(name='nombre').sort_values('nombre', ascending=False).head(5)
    gr['lbl'] = gr['GROUPE RESOLUTION'].apply(lambda x: add_line_breaks(x, max_chars=15))
    fig1 = px.bar(gr, x='lbl', y='nombre', text='nombre', title="Top 5 — Groupe de résolution",
                  color='nombre', color_continuous_scale=px.colors.sequential.Purples, template='plotly_dark')
    fig1.update_traces(textfont_size=12, marker_line_width=0)
    fig1.update_layout(coloraxis_showscale=False, xaxis_tickangle=0)
    apply_layout(fig1, yaxis=dict(visible=False), xaxis=dict(showgrid=False, zeroline=False, color="#8b949e", tickfont=dict(size=11)))
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig1, config={"displayModeBar":False}), style=card_style), width=6))
    ag = data.groupby("AGENCE")["Complaint reference"].count().reset_index(name='nombre').sort_values('nombre').tail(10)
    fig2 = px.bar(ag, y='AGENCE', x='nombre', text='nombre', orientation='h', title='Top 10 agences',
                  color='nombre', color_continuous_scale=px.colors.sequential.Teal, template='plotly_dark')
    fig2.update_traces(textfont_size=12, marker_line_width=0)
    fig2.update_layout(coloraxis_showscale=False)
    apply_layout(fig2, xaxis=dict(visible=False))
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig2, config={"displayModeBar":False}), style=card_style), width=6))
    return graphs

def generate_view_3(data):
    graphs = [dbc.Col(kpi_row(data), width=12)]
    if 'Creator' in data.columns:
        cr = data.groupby("Creator")["Complaint reference"].count().reset_index(name='nombre').sort_values('nombre', ascending=False).head(10)
        cr['lbl'] = cr['Creator'].apply(lambda x: add_line_breaks(x, max_chars=15))
        fig1 = px.bar(cr, x='lbl', y='nombre', text='nombre', title="Top 10 — Par créateur",
                      color='nombre', color_continuous_scale=px.colors.sequential.Reds, template='plotly_dark')
        fig1.update_traces(textfont_size=12, marker_line_width=0)
        fig1.update_layout(coloraxis_showscale=False, xaxis_tickangle=0)
        apply_layout(fig1, yaxis=dict(visible=False), xaxis=dict(showgrid=False, zeroline=False, color="#8b949e", tickfont=dict(size=11)))
        graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig1, config={"displayModeBar":False}), style=card_style), width=12))
    ty = data.groupby("Typology")["Complaint reference"].count().reset_index(name='nombre').sort_values('nombre', ascending=False)
    ty['lbl'] = ty['Typology'].apply(lambda x: add_line_breaks(x, max_chars=15))
    fig2 = px.bar(ty, x='lbl', y='nombre', text='nombre', title="Répartition par typologie",
                  color='nombre', color_continuous_scale=px.colors.sequential.Sunset, template='plotly_dark')
    fig2.update_traces(textfont_size=12, marker_line_width=0)
    fig2.update_layout(coloraxis_showscale=False, xaxis_tickangle=0)
    apply_layout(fig2, yaxis=dict(visible=False), xaxis=dict(showgrid=False, zeroline=False, color="#8b949e", tickfont=dict(size=11)))
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig2, config={"displayModeBar":False}), style=card_style), width=12))
    return graphs

def generate_view_4(data):
    graphs = [dbc.Col(kpi_row(data), width=12)]
    # Donut délai
    dl = data["DELAI_RECLAMATION"].value_counts().reset_index(); dl.columns=["d","c"]
    fig0 = px.pie(dl, values='c', names='d', title="Respect des délais", hole=0.55,
                  color_discrete_map={"PAS HORS DELAI":"#3fb950","HORS DELAI":"#f78166"}, template='plotly_dark')
    fig0.update_traces(textinfo='percent+label')
    apply_layout(fig0)
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig0, config={"displayModeBar":False}), style=card_style), width=6))
    cn = data.groupby("Client notification channel")["Complaint reference"].count().reset_index(name='nombre').sort_values('nombre', ascending=False)
    cn['lbl'] = cn['Client notification channel'].apply(lambda x: add_line_breaks(x, max_chars=15))
    fig1 = px.bar(cn, x='lbl', y='nombre', text='nombre', title="Canaux de notification",
                  color='nombre', color_continuous_scale=px.colors.sequential.Teal, template='plotly_dark')
    fig1.update_traces(textfont_size=12, marker_line_width=0)
    fig1.update_layout(coloraxis_showscale=False, xaxis_tickangle=0)
    apply_layout(fig1, yaxis=dict(visible=False), xaxis=dict(showgrid=False, zeroline=False, color="#8b949e", tickfont=dict(size=11)))
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig1, config={"displayModeBar":False}), style=card_style), width=6))
    ty = data.groupby("Typology")["Complaint reference"].count().reset_index(name='nombre').sort_values('nombre', ascending=False)
    ty['lbl'] = ty['Typology'].apply(lambda x: add_line_breaks(x, max_chars=15))
    fig2 = px.bar(ty, x='lbl', y='nombre', text='nombre', title="Répartition par typologie",
                  color='nombre', color_continuous_scale=px.colors.sequential.Sunset, template='plotly_dark')
    fig2.update_traces(textfont_size=12, marker_line_width=0)
    fig2.update_layout(coloraxis_showscale=False, xaxis_tickangle=0)
    apply_layout(fig2, yaxis=dict(visible=False), xaxis=dict(showgrid=False, zeroline=False, color="#8b949e", tickfont=dict(size=11)))
    graphs.append(dbc.Col(html.Div(dcc.Graph(figure=fig2, config={"displayModeBar":False}), style=card_style), width=12))
    return graphs

def generate_sla_table(data, group_by_field):
    mean_sla = data.pivot_table(index=group_by_field, values='SLA_JOURS', columns=['SLA_ETAPE'], aggfunc='mean').reset_index()
    count_sla = data.pivot_table(index=group_by_field, values='SLA_JOURS', columns=['SLA_ETAPE'], aggfunc='count').reset_index()
    mean_sla_rounded = mean_sla.round(2)
    tables = [
        html.Div(html.H5(f"Moyenne des SLA par {group_by_field}", style={"color":"white","backgroundColor":"#6a0dad","padding":"8px 12px","borderRadius":"8px","fontWeight":"bold","marginTop":"10px","marginBottom":"10px","textAlign":"center"})),
        create_datatable_from_df(mean_sla_rounded, page_size=10),
        html.Div(html.H5(f"Nombre d'entrées SLA par {group_by_field}", style={"color":"white","backgroundColor":"#6a0dad","padding":"8px 12px","borderRadius":"8px","fontWeight":"bold","marginTop":"20px","marginBottom":"10px","textAlign":"center"})),
        create_datatable_from_df(count_sla, page_size=10)
    ]
    return [dbc.Card(tables, style=card_style)]

def generate_sla_ticket_table(data):
    mean_sla = data.pivot_table(index='Complaint reference', values='SLA_JOURS', columns=['SLA_ETAPE'], aggfunc='mean').reset_index()
    mean_sla_rounded = mean_sla.round(2)
    table = [
        html.Div(html.H5("Moyenne des SLA par Ticket", style={"color":"white","backgroundColor":"#6a0dad","padding":"8px 12px","borderRadius":"8px","fontWeight":"bold","marginTop":"10px","marginBottom":"10px","textAlign":"center"})),
        create_datatable_from_df(mean_sla_rounded, page_size=10)
    ]
    return [dbc.Card(table, style=card_style)]

graph_mapping = {
    "btn-1": generate_view_1, "btn-2": generate_view_2,
    "btn-3": generate_view_3, "btn-4": generate_view_4,
    "item-1": lambda data: generate_sla_table(data, 'GROUPE RESOLUTION'),
    "item-2": lambda data: generate_sla_table(data, 'AGENCE'),
    "item-3": generate_sla_ticket_table
}

# ── PAGE ACCUEIL ─────────────────────────────────────────────────────────────
accueil = html.Div([
    html.Div([
        html.Img(src=image_path, style={'height': '80px', 'marginBottom': '24px'}),
        html.H1("Plateforme Réclamations SGCI", style={"color":"white","fontWeight":"800","fontSize":"32px","marginBottom":"12px"}),
        html.P("Choisissez un module pour continuer.", style={"color":"#8b949e","fontSize":"16px","marginBottom":"48px"}),
        html.Div([
            # Bouton 1 — Dashboard
            html.Div([
                html.Div("📊", style={"fontSize":"36px","marginBottom":"12px"}),
                html.H4("Dashboard Réclamations", style={"color":"white","fontWeight":"700","marginBottom":"8px"}),
                html.P("Analysez les réclamations, SLA, agences et segments.", style={"color":"#8b949e","fontSize":"13px","marginBottom":"20px"}),
                dbc.Button("Ouvrir le dashboard", id="btn-open-dashboard", color="primary",
                           style={"borderRadius":"8px","fontWeight":"600","width":"100%"}),
            ], style={"backgroundColor":"#1c2128","border":"1px solid #30363d","borderTop":"3px solid #58a6ff",
                      "borderRadius":"12px","padding":"28px","flex":"1","minWidth":"220px","textAlign":"center"}),
            # Bouton 2 — Lien externe
            html.Div([
                html.Div("🌐", style={"fontSize":"36px","marginBottom":"12px"}),
                html.H4("Ressource Externe", style={"color":"white","fontWeight":"700","marginBottom":"8px"}),
                html.P("Accédez à une ressource de référence externe.", style={"color":"#8b949e","fontSize":"13px","marginBottom":"20px"}),
                html.A(dbc.Button("Ouvrir le lien", color="success",
                           style={"borderRadius":"8px","fontWeight":"600","width":"100%"}),
                       href="https://fr.wikipedia.org/wiki/Lionel_Messi", target="_blank"),
            ], style={"backgroundColor":"#1c2128","border":"1px solid #30363d","borderTop":"3px solid #3fb950",
                      "borderRadius":"12px","padding":"28px","flex":"1","minWidth":"220px","textAlign":"center"}),
            # Bouton 3 — Page future
            html.Div([
                html.Div("🚧", style={"fontSize":"36px","marginBottom":"12px"}),
                html.H4("Module en cours", style={"color":"white","fontWeight":"700","marginBottom":"8px"}),
                html.P("Ce module est en cours de développement.", style={"color":"#8b949e","fontSize":"13px","marginBottom":"20px"}),
                dbc.Button("Bientôt disponible", id="btn-open-module", color="warning",
                           style={"borderRadius":"8px","fontWeight":"600","width":"100%"}),
            ], style={"backgroundColor":"#1c2128","border":"1px solid #30363d","borderTop":"3px solid #ffa657",
                      "borderRadius":"12px","padding":"28px","flex":"1","minWidth":"220px","textAlign":"center"}),
        ], style={"display":"flex","gap":"20px","flexWrap":"wrap","justifyContent":"center"}),
    ], style={"maxWidth":"900px","margin":"0 auto","textAlign":"center"}),
], style={"backgroundColor":"#0d1117","minHeight":"100vh","display":"flex",
          "alignItems":"center","justifyContent":"center","padding":"40px 20px"})

# ── DASHBOARD (original inchangé) ────────────────────────────────────────────
DD_STYLE = {"color": "#000", "fontSize": "13px"}

def filter_block(label, comp):
    return html.Div([html.Label(label, style=label_style), comp,
                     html.Div(style={"marginBottom": "14px"})])

dashboard = dbc.Container(fluid=True,
    style={"backgroundColor": "#0d1117", "minHeight": "100vh", "padding": "0"},
    children=[
        # ── Topbar ──
        html.Div([
            dbc.Row([
                dbc.Col(html.Img(src=image_path, style={"height": "38px"}), width="auto"),
                dbc.Col(html.Span("Dashboard Réclamations", style={
                    "color": "#e6edf3", "fontWeight": "700", "fontSize": "17px", "lineHeight": "38px",
                }), width="auto"),
                dbc.Col(
                    dbc.Button("← Accueil", id="btn-retour-accueil", color="secondary", size="sm",
                               style={**btn_style, "fontSize": "12px"}),
                    width="auto", style={"marginLeft": "auto"}
                ),
            ], align="center"),
        ], style={
            "backgroundColor": "#161b22", "borderBottom": "1px solid #30363d",
            "padding": "12px 24px", "position": "sticky", "top": "0", "zIndex": "100",
        }),
        dbc.Row([
            # ── Sidebar filtres ──
            dbc.Col([
                html.Div([
                    html.Div([
                        html.Span("⚙ ", style={"opacity": "0.7"}),
                        html.Span("Filtres", style={"color": "#e6edf3", "fontWeight": "700",
                                                     "fontSize": "13px", "textTransform": "uppercase", "letterSpacing": "1px"}),
                    ], style={"marginBottom": "18px", "paddingBottom": "12px", "borderBottom": "1px solid #30363d"}),
                    filter_block("🎫 Ticket",     dcc.Dropdown(id="ticket-filter",      clearable=True, style=DD_STYLE)),
                    filter_block("🏢 Agence",     dcc.Dropdown(id="agence-filter",      multi=True,     style=DD_STYLE)),
                    filter_block("📅 Année",      dcc.Dropdown(id="annee-dropdown",     clearable=True, style=DD_STYLE)),
                    filter_block("🗓 Mois",       dcc.Dropdown(id="mois-dropdown",      clearable=True, style=DD_STYLE)),
                    filter_block("👤 Segmentation",dcc.Dropdown(id="segmentation-filter",clearable=True,style=DD_STYLE)),
                    filter_block("⏱ Délai",      dcc.Dropdown(id="delai-filter",       multi=True,     style=DD_STYLE)),
                    dbc.Button([html.I(className="bi bi-arrow-counterclockwise me-2"), "Réinitialiser"],
                               id="reset-filters", color="danger", outline=True, size="sm",
                               className="w-100 mt-1", style={**btn_style, "fontSize": "12px"}),
                    dcc.Upload(id="upload-data",
                        children=html.Div([
                            html.Span("☁", style={"fontSize": "26px", "display": "block", "marginBottom": "4px"}),
                            html.Span("Glisser / cliquer", style={"fontSize": "12px", "color": "#e6edf3", "fontWeight": "600"}),
                            html.Br(),
                            html.Span("CSV ou Excel", style={"fontSize": "11px", "color": "#8b949e"}),
                        ], style={"textAlign": "center"}),
                        style={"marginTop": "20px", "borderWidth": "1.5px", "borderStyle": "dashed",
                               "borderColor": "#58a6ff", "borderRadius": "10px", "padding": "18px 8px", "cursor": "pointer",
                               "backgroundColor": "rgba(88,166,255,0.04)"},
                        max_size=50 * 1024 * 1024,
                        multiple=False),
                    html.Div(id='upload-status'),
                ], style={
                    "backgroundColor": "#161b22", "borderRight": "1px solid #30363d",
                    "padding": "20px 16px", "minHeight": "calc(100vh - 64px)",
                }),
            ], width=2, style={"padding": "0"}),
            # ── Zone principale ──
            dbc.Col([
                html.Div([
                    dbc.Row([
                        dbc.Col(dbc.Button("📊 Vue Générale", id="btn-1", color="primary",   className="w-100", style={**btn_style, "fontSize": "12px", "padding": "8px"}), width=2),
                        dbc.Col(dbc.Button("👥 Par Groupe",   id="btn-2", color="secondary", className="w-100", style={**btn_style, "fontSize": "12px", "padding": "8px"}), width=2),
                        dbc.Col(dbc.Button("🎫 Par Ticket",  id="btn-3", color="secondary", className="w-100", style={**btn_style, "fontSize": "12px", "padding": "8px"}), width=2),
                        dbc.Col(dbc.Button("📡 Canaux",      id="btn-4", color="secondary", className="w-100", style={**btn_style, "fontSize": "12px", "padding": "8px"}), width=2),
                        dbc.Col(dbc.DropdownMenu(label="📋 SLA PAR", children=items, id="sla-dropdown",
                                                  color="info", className="w-100"), width=2),
                    ], className="mb-4 g-2"),
                    html.Div(id="content", style={"minHeight": "400px"}),
                    dcc.Store(id="stored-data"),
                    dcc.Store(id="stored-active-view"),
                    # ── Panneau debug (toujours visible en production) ──
                    html.Details([
                        html.Summary("🔍 Logs de débogage", style={
                            "color": "#8b949e", "fontSize": "12px", "cursor": "pointer",
                            "padding": "8px 0", "userSelect": "none",
                        }),
                        html.Div([
                            dbc.Button("⚡ Tester les callbacks", id="btn-test-cb", size="sm",
                                color="secondary", outline=True,
                                style={"fontSize": "11px", "marginBottom": "8px"}),
                            html.Div(id="debug-panel", style={
                                "backgroundColor": "#0d1117",
                                "border": "1px solid #30363d",
                                "borderRadius": "8px",
                                "padding": "12px",
                                "fontFamily": "monospace",
                                "fontSize": "12px",
                                "color": "#e6edf3",
                                "whiteSpace": "pre-wrap",
                                "maxHeight": "300px",
                                "overflowY": "auto",
                            }),
                        ]),
                    ], style={"marginTop": "32px", "borderTop": "1px solid #30363d", "paddingTop": "16px"}),
                ], style={"padding": "24px"}),
            ], width=10, style={"padding": "0"}),
        ], style={"margin": "0"}),
    ]
)

# ═══════════════════════════════════════════════════════════════
# MODULE KPI — à coller dans reclam_final.py
# Remplace le bloc "MODULE FUTUR" et ajoute les callbacks
# ═══════════════════════════════════════════════════════════════
import sqlite3, json
from dash import dash_table

DB_PATH = "kpi_historique.db"

# ── Mapping UC → Agences (référentiel fixe) ──────────────────
UC_MAPPING = {
    "UC-ABIDJAN EST DEUX PLATEAUX": [
        "AGHIEN", "COCODY VALLONS", "LYCEE TECHNIQUE",
        "AGENCE COCODY CENTRE", "COCODY 2 PLATEAUX",
        "RUE DES JARDINS", "ANGRE DJIBI CENTRE",
    ],
    "UC-ABIDJAN CENTRE": [
        "PYRAMIDE", "AKWABA", "LONGCHAMPS", "DU PARC",
        "COMMERCE", "PRESTIGE", "PLATEAU SIEGE",
        "PRIVILEGE", "CITE FINANCIERE",
    ],
    "UC-ABIDJAN EST RIVIERA": [
        "BINGERVILLE", "ABATTA", "RIVIERA PALMERAIE",
        "RIVIERA GOLF", "RIVIERA ANONO", "RIVIERA SAINTE FAMILLE",
    ],
    "UC-ABIDJAN NORD": [
        "INDENIE", "WILLIAMSVILLE", "ADJAME MARCHE", "ADJAME LIBERTE",
        "AGBOVILLE", "ABOBO", "ABOBO SAMAKE", "ANYAMA", "PLATEAU-DOKUI",
    ],
    "UC-ABIDJAN OUEST": [
        "SONGON", "DABOU", "YOPOUGON ZONE INDUSTRIELLE",
        "YOPOUGON ANANERAIE", "YOPOUGON BEL AIR", "YOPOUGON SAINT ANDRE",
        "YOPOUGON NIANGON SUD", "YOPOUGON NIANGON NORD", "YOPOUGON FIGAYO",
    ],
    "UC-ABIDJAN SUD": [
        "ELITE", "MOSQUEE", "MARINE", "BIETRY",
        "AGENCE TOTAL", "AUTOROUTE",
    ],
}

# Index inverse : agence → UC
AGENCE_TO_UC = {ag: uc for uc, ags in UC_MAPPING.items() for ag in ags}

# ── Init SQLite ──────────────────────────────────────────────
def init_db():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    # Table CRC : une ligne par indicateur par mois
    cur.execute("""CREATE TABLE IF NOT EXISTS kpi_crc (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mois TEXT, indicateur TEXT, objectif TEXT, valeur TEXT,
        UNIQUE(mois, indicateur)
    )""")
    # Table SAT : une ligne par entité par feuille par mois
    cur.execute("""CREATE TABLE IF NOT EXISTS kpi_sat (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mois TEXT, feuille TEXT, entite TEXT, valeur TEXT,
        UNIQUE(mois, feuille, entite)
    )""")
    # Table SAT v2 : avec colonnes uc et is_uc
    cur.execute("""CREATE TABLE IF NOT EXISTS kpi_sat_v2 (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mois TEXT, feuille TEXT, entite TEXT, valeur TEXT, uc TEXT, is_uc INTEGER,
        UNIQUE(mois, feuille, entite)
    )""")
    # Table SAT PRO : même structure que kpi_sat_v2
    cur.execute("""CREATE TABLE IF NOT EXISTS kpi_sat_pro_v2 (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mois TEXT, feuille TEXT, entite TEXT, valeur TEXT, uc TEXT, is_uc INTEGER,
        UNIQUE(mois, feuille, entite)
    )""")
    # Table Feuille1 : structure Segment × Indicateur × Mois
    cur.execute("""CREATE TABLE IF NOT EXISTS kpi_feuille1 (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source TEXT, mois TEXT, segment TEXT, indicateur TEXT, valeur TEXT,
        UNIQUE(source, mois, segment, indicateur)
    )""")
    # mapping UC géré par UC_MAPPING statique
    con.commit(); con.close()

init_db()

def get_uc_list():
    return sorted(UC_MAPPING.keys())

def get_agences_for_uc(uc=None):
    if uc:
        return UC_MAPPING.get(uc, [])
    return sorted(AGENCE_TO_UC.keys())

# ── Parsers ──────────────────────────────────────────────────
def parse_crc(contents, filename, mois):
    """KPI CRC.xlsx : col A=indicateur, col B=objectif, col C+=mois.
    Recherche robuste de la ligne d'en-tête (gère les cellules fusionnées)."""
    _, data = contents.split(',', 1)
    raw = base64.b64decode(data)
    rows = []
    try:
        xf = pd.read_excel(io.BytesIO(raw), sheet_name=0, header=None, engine='openpyxl')
        if xf.empty or xf.shape[0] < 2:
            return False, "Fichier vide ou structure invalide."

        # Recherche de la ligne d'en-tête : celle qui contient le plus de mots-clés mois
        mois_kw = ['JANV','FEV','FÉV','MARS','AVR','MAI','JUIN','JUIL','AOUT','AOÛT',
                   'SEPT','OCT','NOV','DEC','DÉC']
        header_row = None
        best_score = 0
        for i in range(min(10, len(xf))):
            row_str = ' '.join([str(v).upper() for v in xf.iloc[i] if str(v).strip() not in ('nan','')])
            score = sum(1 for k in mois_kw if k in row_str)
            if score > best_score:
                best_score = score
                header_row = i
        if header_row is None:
            header_row = 2  # fallback raisonnable

        data_start = header_row + 1
        for i in range(data_start, len(xf)):
            indic = str(xf.iloc[i, 1]).strip()
            obj   = str(xf.iloc[i, 2]).strip() if xf.shape[1] > 2 else ''
            if indic in ('nan', '') or indic.upper() == 'NAN':
                continue
            vals_row = xf.iloc[i, 3:].dropna() if xf.shape[1] > 3 else pd.Series(dtype=object)
            valeur = str(vals_row.iloc[-1]).strip() if len(vals_row) > 0 else 'N/A'
            rows.append((mois, indic, obj if obj != 'nan' else '', valeur))

        if not rows:
            return False, "Aucun indicateur trouvé — vérifiez la structure du fichier."
    except Exception as e:
        return False, f"Erreur lecture : {str(e)}"

    try:
        con = sqlite3.connect(DB_PATH)
        con.executemany(
            "INSERT OR REPLACE INTO kpi_crc (mois, indicateur, objectif, valeur) VALUES (?,?,?,?)",
            rows
        )
        con.commit(); con.close()
        return True, f"{len(rows)} indicateurs CRC sauvegardés pour {mois}"
    except Exception as e:
        return False, f"Erreur SQLite : {str(e)}"




def parse_feuille1(df, sheet_name, mois, source, con):
    """Parse la Feuille1 : col A=Segment (fusionné), col B=Indicateur, col C+=mois."""
    rows = []
    mois_kw = ['JANV','FEV','FÉV','MARS','AVR','MAI','JUIN','JUIL',
               'AOUT','AOÛT','SEPT','OCT','NOV','DEC','DÉC']

    # Trouver la ligne d'en-tête des mois
    hrow = 1
    for i in range(min(10, len(df))):
        row_str = ' '.join([str(v).upper() for v in df.iloc[i] if str(v).strip() != 'nan'])
        if any(k in row_str for k in mois_kw):
            hrow = i
            break

    # Extraire les labels de mois (col C+)
    mois_labels = []
    for j in range(2, df.shape[1]):
        v = str(df.iloc[hrow, j]).strip()
        if v and v.upper() != 'NAN':
            mois_labels.append((j, v))

    # Lire les données — propager le segment (colonne A fusionnée)
    current_segment = None
    for i in range(hrow + 1, len(df)):
        seg_val   = str(df.iloc[i, 0]).strip()
        indic_val = str(df.iloc[i, 1]).strip()

        if seg_val and seg_val.upper() not in ('NAN', ''):
            current_segment = seg_val.upper()
        if not indic_val or indic_val.upper() in ('NAN', '') or not current_segment:
            continue
        # Ignorer les lignes séparateurs noirs (toutes NaN sauf col A/B)
        all_vals = [str(df.iloc[i, j]).strip() for j in range(2, df.shape[1])]
        if all(v in ('nan', '', 'NAN') for v in all_vals):
            continue

        for j, m_label in mois_labels:
            v = str(df.iloc[i, j]).strip()
            valeur = '' if v in ('nan', 'NAN', '') else v
            rows.append((source, mois, current_segment, indic_val, valeur))

    if rows:
        con.executemany(
            "INSERT OR REPLACE INTO kpi_feuille1 (source, mois, segment, indicateur, valeur) VALUES (?,?,?,?,?)",
            rows
        )
    return len(rows)


def get_feuille1_history(source, mois_filter=None):
    con = sqlite3.connect(DB_PATH)
    try:
        q = "SELECT * FROM kpi_feuille1 WHERE source=?"
        params = [source]
        if mois_filter:
            q += " AND mois=?"; params.append(mois_filter)
        q += " ORDER BY mois, segment, indicateur"
        return pd.read_sql(q, con, params=params)
    except: return pd.DataFrame()
    finally: con.close()


def render_feuille1(source, mois_filter):
    """Affiche la Feuille1 : tableau Segment × Indicateur × Mois."""
    df = get_feuille1_history(source, mois_filter)
    if df.empty:
        return None  # pas de données feuille1, on n'affiche rien

    segments   = df["segment"].unique().tolist()
    indicateurs = df["indicateur"].unique().tolist()
    mois_list  = sorted(df["mois"].unique().tolist())

    SEG_COLORS = {"PRI": "#58a6ff", "PRO": "#3fb950", "CORPO": "#f78166"}
    INDIC_BG   = {"Indice SAT": "rgba(255,166,87,0.15)",
                  "NPS":        "rgba(88,166,255,0.15)",
                  "Nombre de repondants": "rgba(63,185,80,0.10)"}

    # ── En-tête ──
    header_cells = [
        html.Th("SEGMENT",    style={"backgroundColor":"#1F497D","color":"white","padding":"10px 14px",
                                      "fontWeight":"700","fontSize":"11px","border":"1px solid #30363d",
                                      "minWidth":"100px","textAlign":"center"}),
        html.Th("INDICATEUR", style={"backgroundColor":"#1F497D","color":"white","padding":"10px 14px",
                                      "fontWeight":"700","fontSize":"11px","border":"1px solid #30363d",
                                      "minWidth":"180px","textAlign":"left"}),
    ]
    for m in mois_list:
        header_cells.append(html.Th(m, style={
            "backgroundColor":"#1F497D","color":"white","padding":"10px 14px",
            "fontWeight":"700","fontSize":"11px","border":"1px solid #30363d",
            "minWidth":"110px","textAlign":"center",
        }))

    # ── Corps ──
    body_rows = []
    for seg in segments:
        color = SEG_COLORS.get(seg.upper(), "#e6edf3")
        df_seg = df[df["segment"] == seg]
        indics = df_seg["indicateur"].unique().tolist()
        for k, indic in enumerate(indics):
            df_ind = df_seg[df_seg["indicateur"] == indic]
            bg_indic = INDIC_BG.get(indic, "transparent")
            cells = []
            # Colonne Segment (seulement sur la première ligne du groupe)
            if k == 0:
                cells.append(html.Td(seg, rowSpan=len(indics), style={
                    "backgroundColor": "#1c2128",
                    "color": color, "fontWeight": "800", "fontSize": "13px",
                    "textAlign": "center", "verticalAlign": "middle",
                    "padding": "10px 14px", "border": "1px solid #30363d",
                    "letterSpacing": "1px",
                }))
            # Colonne Indicateur
            cells.append(html.Td(indic, style={
                "backgroundColor": bg_indic, "color": "#e6edf3",
                "fontWeight": "600", "fontSize": "12px",
                "padding": "9px 14px", "border": "1px solid #30363d",
                "textAlign": "left",
            }))
            # Valeurs par mois
            for m in mois_list:
                row_m = df_ind[df_ind["mois"] == m]
                val = row_m["valeur"].values[0] if not row_m.empty else ""
                val = "" if pd.isna(val) or str(val) in ("nan","None","") else str(val)
                cells.append(html.Td(val, style={
                    "backgroundColor": bg_indic, "color": color if val else "#8b949e",
                    "fontWeight": "700" if val else "400",
                    "fontSize": "12px", "padding": "9px 14px",
                    "border": "1px solid #30363d", "textAlign": "center",
                }))
            body_rows.append(html.Tr(cells))

        # Séparateur entre segments
        if seg != segments[-1]:
            body_rows.append(html.Tr([
                html.Td(colSpan=len(mois_list)+2, style={
                    "backgroundColor": "#000", "height": "4px",
                    "border": "none", "padding": "0",
                })
            ]))

    return html.Div([
        html.H6("Enquête à froid — par segment", style={
            "color":"#e6edf3","fontWeight":"700","marginBottom":"12px"
        }),
        html.Div(
            html.Table(
                [html.Thead(html.Tr(header_cells)), html.Tbody(body_rows)],
                style={"width":"100%","borderCollapse":"collapse",
                       "fontFamily":"'Segoe UI', sans-serif"},
            ),
            style={"overflowX":"auto","border":"1px solid #30363d","borderRadius":"8px"},
        ),
    ], style={**card_style, "marginBottom":"20px"})


def parse_sat(contents, filename, mois):
    """KPIS SAT.xlsx : plusieurs feuilles avec structure UC > Agences.
    Le mapping UC → Agences est géré par UC_MAPPING (référentiel statique)."""
    _, data = contents.split(',', 1)
    raw = base64.b64decode(data)
    rows = []

    try:
        xl = pd.ExcelFile(io.BytesIO(raw), engine='openpyxl')
        mois_kw = ['JANV','FEV','FÉV','MARS','AVR','MAI','JUIN','JUIL',
                   'AOUT','AOÛT','SEPT','OCT','NOV','DEC','DÉC']

        con = sqlite3.connect(DB_PATH)
        nb_f1 = 0
        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name, header=None)
            if df.empty: continue

            # Feuille1 → parser dédié (col A=segment, col B=indicateur)
            if sheet_name.strip().lower() in ("feuil1","feuille1","sheet1","feuil 1"):
                nb_f1 = parse_feuille1(df, sheet_name, mois, "sat", con)
                continue

            # Autres feuilles → structure UC > Agences
            hrow = 1
            for i in range(min(10, len(df))):
                row_str = ' '.join([str(v).upper() for v in df.iloc[i] if str(v).strip() != 'nan'])
                if any(k in row_str for k in mois_kw):
                    hrow = i; break

            for i in range(hrow + 1, len(df)):
                entite = str(df.iloc[i, 0]).strip().upper()
                if not entite or entite in ('NAN', ''): continue
                vals = df.iloc[i, 1:].dropna()
                valeur = str(vals.iloc[-1]).strip() if len(vals) > 0 else 'N/A'
                is_uc  = 1 if entite in UC_MAPPING else 0
                uc_ref = entite if is_uc else AGENCE_TO_UC.get(entite, None)
                rows.append((mois, sheet_name, entite, valeur, uc_ref, is_uc))

        con.commit(); con.close()

    except Exception as e:
        return False, f"Erreur lecture : {str(e)}"

    if not rows and nb_f1 == 0:
        return False, "Aucune ligne trouvée dans le fichier."

    try:
        con = sqlite3.connect(DB_PATH)
        con.executemany(
            "INSERT OR REPLACE INTO kpi_sat_v2 (mois, feuille, entite, valeur, uc, is_uc) VALUES (?,?,?,?,?,?)",
            rows
        )
        con.commit(); con.close()
        nb_uc = sum(1 for r in rows if r[5] == 1)
        nb_ag = sum(1 for r in rows if r[5] == 0)
        return True, f"{len(rows)} lignes SAT ({nb_uc} UCs, {nb_ag} agences) + {nb_f1} lignes Feuille1"
    except Exception as e:
        return False, f"Erreur SQLite : {str(e)}"

def get_sat_history(feuille=None, uc=None, agences=None):
    """Lecture avec filtres UC et agences."""
    con = sqlite3.connect(DB_PATH)
    try:
        # Essayer la table enrichie v2 d'abord
        tables = [r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        table = "kpi_sat_v2" if "kpi_sat_v2" in tables else "kpi_sat"
        q = f"SELECT * FROM {table} WHERE 1=1"
        params = []
        if feuille:
            q += " AND feuille=?"; params.append(feuille)
        if uc:
            q += " AND uc=?"; params.append(uc)
        if agences:
            placeholders = ','.join(['?'] * len(agences))
            q += f" AND entite IN ({placeholders})"; params.extend(agences)
        q += " ORDER BY mois"
        df = pd.read_sql(q, con, params=params)
        return df
    except: return pd.DataFrame()
    finally: con.close()


def parse_sat_pro(contents, filename, mois):
    """KPIS SAT PRO.xlsx : même structure que KPIS SAT.xlsx, même parser."""
    _, data = contents.split(',', 1)
    raw = base64.b64decode(data)
    rows = []
    try:
        xl = pd.ExcelFile(io.BytesIO(raw), engine='openpyxl')
        mois_kw = ['JANV','FEV','FÉV','MARS','AVR','MAI','JUIN','JUIL',
                   'AOUT','AOÛT','SEPT','OCT','NOV','DEC','DÉC']
        con = sqlite3.connect(DB_PATH)
        nb_f1 = 0
        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name, header=None)
            if df.empty: continue

            # Feuille1 → parser dédié
            if sheet_name.strip().lower() in ("feuil1","feuille1","sheet1","feuil 1"):
                nb_f1 = parse_feuille1(df, sheet_name, mois, "sat_pro", con)
                continue

            hrow = 1
            for i in range(min(10, len(df))):
                row_str = ' '.join([str(v).upper() for v in df.iloc[i] if str(v).strip() != 'nan'])
                if any(k in row_str for k in mois_kw):
                    hrow = i; break
            for i in range(hrow + 1, len(df)):
                entite = str(df.iloc[i, 0]).strip().upper()
                if not entite or entite in ('NAN', ''): continue
                vals = df.iloc[i, 1:].dropna()
                valeur = str(vals.iloc[-1]).strip() if len(vals) > 0 else 'N/A'
                is_uc  = 1 if entite in UC_MAPPING else 0
                uc_ref = entite if is_uc else AGENCE_TO_UC.get(entite, None)
                rows.append((mois, sheet_name, entite, valeur, uc_ref, is_uc))

        con.commit(); con.close()

    except Exception as e:
        return False, f"Erreur lecture : {str(e)}"

    if not rows and nb_f1 == 0:
        return False, "Aucune ligne trouvée dans le fichier."
    try:
        con = sqlite3.connect(DB_PATH)
        con.executemany(
            "INSERT OR REPLACE INTO kpi_sat_pro_v2 (mois, feuille, entite, valeur, uc, is_uc) VALUES (?,?,?,?,?,?)",
            rows
        )
        con.commit(); con.close()
        nb_uc = sum(1 for r in rows if r[5] == 1)
        nb_ag = sum(1 for r in rows if r[5] == 0)
        return True, f"{len(rows)} lignes SAT PRO ({nb_uc} UCs, {nb_ag} agences) + {nb_f1} lignes Feuille1"
    except Exception as e:
        return False, f"Erreur SQLite : {str(e)}"


# ── Lecture historique ───────────────────────────────────────
def get_crc_history():
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT * FROM kpi_crc ORDER BY mois", con)
    con.close(); return df



def get_sat_pro_history(uc=None, agences=None):
    con = sqlite3.connect(DB_PATH)
    try:
        q = "SELECT * FROM kpi_sat_pro_v2 WHERE 1=1"
        params = []
        if uc:
            q += " AND uc=?"; params.append(uc)
        if agences:
            placeholders = ','.join(['?'] * len(agences))
            q += f" AND entite IN ({placeholders})"; params.extend(agences)
        q += " ORDER BY mois"
        return pd.read_sql(q, con, params=params)
    except: return pd.DataFrame()
    finally: con.close()

def get_mois_list():
    con = sqlite3.connect(DB_PATH)
    mois = []
    for table in ['kpi_crc','kpi_sat_v2','kpi_sat_pro_v2']:
        try:
            df = pd.read_sql(f"SELECT DISTINCT mois FROM {table}", con)
            mois.extend(df['mois'].tolist())
        except: pass
    con.close()
    return sorted(list(set(mois)))

# ── Layout module ────────────────────────────────────────────
def make_status_badge(text, color):
    return html.Span(text, style={
        "backgroundColor": color, "color": "#fff", "borderRadius": "6px",
        "padding": "3px 10px", "fontSize": "11px", "fontWeight": "700",
        "marginLeft": "8px",
    })

def upload_zone(label, uid, icon):
    return html.Div([
        html.Label([icon, f"  {label}"], style={**label_style, "fontSize": "12px", "marginBottom": "6px"}),
        dcc.Upload(id=uid,
            children=html.Div([
                html.Span("☁", style={"fontSize": "20px", "display": "block", "marginBottom": "2px"}),
                html.Span("Choisir le fichier", style={"fontSize": "11px", "color": "#8b949e"}),
            ], style={"textAlign": "center"}),
            style={"borderWidth": "1.5px", "borderStyle": "dashed", "borderColor": "#30363d",
                   "borderRadius": "8px", "padding": "10px 6px", "cursor": "pointer",
                   "backgroundColor": "#0d1117"},
            max_size=50 * 1024 * 1024,
            multiple=False),
        html.Div(id=f"{uid}-status", style={"fontSize": "11px", "marginTop": "4px", "minHeight": "16px"}),
        html.Div(style={"marginBottom": "14px"}),
    ])

MOIS_OPTIONS = [
    {"label": m, "value": m} for m in
    ["Janvier","Février","Mars","Avril","Mai","Juin",
     "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
]
ANNEE_OPTIONS = [{"label": str(a), "value": str(a)} for a in range(2024, 2028)]

module = dbc.Container(fluid=True,
    style={"backgroundColor": "#0d1117", "minHeight": "100vh", "padding": "0"},
    children=[
        # Topbar
        html.Div([dbc.Row([
            dbc.Col(html.Img(src='assets/logo.png', style={"height": "36px"}), width="auto"),
            dbc.Col(html.Span("Tableau de Bord KPI", style={
                "color": "#e6edf3", "fontWeight": "700", "fontSize": "16px",
                "fontFamily": "'Syne', sans-serif", "lineHeight": "36px"}), width="auto"),
            dbc.Col(html.Button("← Accueil", id="btn-retour-accueil-module", n_clicks=0, style={
                **btn_style, "backgroundColor": "transparent", "border": "1px solid #30363d",
                "color": "#8b949e", "padding": "6px 16px", "fontSize": "12px"}),
                width="auto", style={"marginLeft": "auto"}),
        ], align="center")],
        style={"backgroundColor": "#161b22", "borderBottom": "1px solid #30363d",
               "padding": "12px 24px", "position": "sticky", "top": "0", "zIndex": "100"}),

        dbc.Row([
            # ── Sidebar ──
            dbc.Col([html.Div([
                html.Div([
                    html.Span("📥 ", style={"opacity": "0.8"}),
                    html.Span("Import mensuel", style={"color": "#e6edf3", "fontWeight": "700",
                                                        "fontSize": "13px", "textTransform": "uppercase",
                                                        "letterSpacing": "1px"}),
                ], style={"marginBottom": "16px", "paddingBottom": "12px", "borderBottom": "1px solid #30363d"}),

                # Sélection mois/année
                html.Div([
                    html.Label("📅 Mois de référence", style=label_style),
                    dcc.Dropdown(id="kpi-mois-select", options=MOIS_OPTIONS, clearable=False,
                                 placeholder="Mois...", style={"color": "#000", "fontSize": "12px",
                                                               "marginBottom": "6px"}),
                    dcc.Dropdown(id="kpi-annee-select", options=ANNEE_OPTIONS, clearable=False,
                                 value="2026", style={"color": "#000", "fontSize": "12px"}),
                    html.Div(style={"marginBottom": "16px"}),
                ]),

                upload_zone("KPI CRC",     "upload-crc",     "📞"),
                upload_zone("KPIS SAT",    "upload-sat",     "😊"),
                upload_zone("KPIS SAT PRO","upload-sat-pro", "⭐"),

                html.Hr(style={"borderColor": "#30363d", "marginTop": "8px"}),

                # Filtre historique
                html.Label("🗂 Filtrer par mois", style=label_style),
                dcc.Dropdown(id="kpi-hist-filter", clearable=True,
                             placeholder="Tous les mois...",
                             style={"color": "#000", "fontSize": "12px", "marginBottom": "14px"}),

                # Filtres UC / Agence (visibles seulement pour onglet SAT)
                html.Div(id="kpi-uc-filters", children=[
                    html.Label("🏢 Filtrer par UC", style=label_style),
                    dcc.Dropdown(id="kpi-uc-filter", clearable=True,
                                 placeholder="Toutes les UCs...",
                                 style={"color": "#000", "fontSize": "12px", "marginBottom": "8px"}),
                    html.Label("📍 Filtrer par Agence", style=label_style),
                    dcc.Dropdown(id="kpi-agence-filter", clearable=True, multi=True,
                                 placeholder="Toutes les agences...",
                                 style={"color": "#000", "fontSize": "12px", "marginBottom": "14px"}),
                ]),

            ], style={"backgroundColor": "#161b22", "borderRight": "1px solid #30363d",
                      "padding": "20px 16px", "minHeight": "calc(100vh - 64px)"})],
            width=2, style={"padding": "0"}),

            # ── Zone principale ──
            dbc.Col([html.Div([
                dbc.Row([
                    dbc.Col(dbc.Tabs(id="kpi-tabs", active_tab="tab-crc", children=[
                        dbc.Tab(label="📞 KPI CRC",         tab_id="tab-crc"),
                        dbc.Tab(label="😊 Satisfaction",    tab_id="tab-sat"),
                        dbc.Tab(label="⭐ SAT PRO",      tab_id="tab-sat-pro"),
                    ]), width=9),
                    dbc.Col(
                        dbc.Button([html.I(className="bi bi-download me-2"), "Exporter Excel"],
                            id="btn-export-kpi", color="success", outline=True, size="sm",
                            className="w-100", style={**btn_style, "fontSize": "12px"}),
                        width=3,
                    ),
                ], align="center", className="mb-3"),
                dcc.Download(id="download-kpi-excel"),
                html.Div(id="kpi-tab-content", style={"minHeight": "400px"}),
            ], style={"padding": "24px"})],
            width=10, style={"padding": "0"}),
        ], style={"margin": "0"}),
    ]
)


# ═══════════════════════════════════════════════════════
# CALLBACKS MODULE KPI
# ═══════════════════════════════════════════════════════

def register_kpi_callbacks(app):

    # ── Upload CRC ──────────────────────────────────────
    @app.callback(
        Output("upload-crc-status", "children"),
        Output("kpi-hist-filter", "options"),
        Output("upload-crc", "contents"),
        Input("upload-crc", "contents"),
        State("upload-crc", "filename"),
        State("kpi-mois-select", "value"),
        State("kpi-annee-select", "value"),
        prevent_initial_call=True,
    )
    def upload_crc(contents, filename, mois, annee):
        if not contents:
            return "", dash.no_update, dash.no_update
        if not mois or not annee:
            return html.Div("⚠ Sélectionnez d'abord le mois et l'année !",
                style={"color": "#ffa657", "fontWeight": "700", "padding": "6px",
                       "backgroundColor": "rgba(255,166,87,0.1)", "borderRadius": "6px"}), dash.no_update, None
        mois_str = f"{mois} {annee}"
        ok, msg = parse_crc(contents, filename, mois_str)
        color = "#3fb950" if ok else "#f78166"
        icon  = "✅" if ok else "❌"
        opts  = [{"label": m, "value": m} for m in get_mois_list()]
        return html.Span(f"{icon} {msg}", style={"color": color}), opts, None

    # ── Upload SAT ──────────────────────────────────────
    @app.callback(
        Output("upload-sat-status", "children"),
        Output("kpi-hist-filter", "options", allow_duplicate=True),
        Output("upload-sat", "contents"),
        Input("upload-sat", "contents"),
        State("upload-sat", "filename"),
        State("kpi-mois-select", "value"),
        State("kpi-annee-select", "value"),
        prevent_initial_call=True,
    )
    def upload_sat(contents, filename, mois, annee):
        if not contents:
            return "", dash.no_update, dash.no_update
        if not mois or not annee:
            return html.Div("⚠ Sélectionnez d'abord le mois et l'année !",
                style={"color": "#ffa657", "fontWeight": "700", "padding": "6px",
                       "backgroundColor": "rgba(255,166,87,0.1)", "borderRadius": "6px"}), dash.no_update, None
        mois_str = f"{mois} {annee}"
        ok, msg = parse_sat(contents, filename, mois_str)
        color = "#3fb950" if ok else "#f78166"
        icon  = "✅" if ok else "❌"
        opts  = [{"label": m, "value": m} for m in get_mois_list()]
        return html.Span(f"{icon} {msg}", style={"color": color}), opts, None

    # ── Upload SAT PRO ───────────────────────────────────
    @app.callback(
        Output("upload-sat-pro-status", "children"),
        Output("kpi-hist-filter", "options", allow_duplicate=True),
        Output("upload-sat-pro", "contents"),
        Input("upload-sat-pro", "contents"),
        State("upload-sat-pro", "filename"),
        State("kpi-mois-select", "value"),
        State("kpi-annee-select", "value"),
        prevent_initial_call=True,
    )
    def upload_sat_pro(contents, filename, mois, annee):
        if not contents:
            return "", dash.no_update, dash.no_update
        if not mois or not annee:
            return html.Div("⚠ Sélectionnez d'abord le mois et l'année !",
                style={"color": "#ffa657", "fontWeight": "700", "padding": "6px",
                       "backgroundColor": "rgba(255,166,87,0.1)", "borderRadius": "6px"}), dash.no_update, None
        mois_str = f"{mois} {annee}"
        ok, msg = parse_sat_pro(contents, filename, mois_str)
        color = "#3fb950" if ok else "#f78166"
        icon  = "✅" if ok else "❌"
        opts  = [{"label": m, "value": m} for m in get_mois_list()]
        return html.Span(f"{icon} {msg}", style={"color": color}), opts, None

    # ── Alimenter le filtre UC ──────────────────────────
    @app.callback(
        Output("kpi-uc-filter", "options"),
        Input("kpi-tabs", "active_tab"),
        Input("upload-sat", "contents"),
        Input("upload-sat-pro", "contents"),
    )
    def refresh_uc_options(tab, _1, _2):
        ucs = get_uc_list()
        return [{"label": uc, "value": uc} for uc in ucs]

    # ── Filtre agences en cascade selon UC ──────────────
    @app.callback(
        Output("kpi-agence-filter", "options"),
        Output("kpi-agence-filter", "value"),
        Input("kpi-uc-filter", "value"),
        Input("upload-sat", "contents"),
        Input("upload-sat-pro", "contents"),
    )
    def refresh_agence_options(uc, _1, _2):
        agences = get_agences_for_uc(uc)
        opts = [{"label": a, "value": a} for a in agences]
        return opts, None

    # ── Afficher/masquer les filtres UC selon l'onglet ──
    @app.callback(
        Output("kpi-uc-filters", "style"),
        Input("kpi-tabs", "active_tab"),
    )
    def toggle_uc_filters(tab):
        if tab in ("tab-sat", "tab-sat-pro"):
            return {"display": "block"}
        return {"display": "none"}

    # ── Rendu des onglets ────────────────────────────────
    @app.callback(
        Output("kpi-tab-content", "children"),
        Input("kpi-tabs", "active_tab"),
        Input("kpi-hist-filter", "value"),
        Input("kpi-uc-filter", "value"),
        Input("kpi-agence-filter", "value"),
        Input("upload-crc", "contents"),
        Input("upload-sat", "contents"),
        Input("upload-sat-pro", "contents"),
    )
    def render_tab(tab, mois_filter, uc_filter, agence_filter, *_):
        if tab == "tab-crc":
            return render_crc(mois_filter)
        elif tab == "tab-sat":
            return render_sat(mois_filter, uc_filter, agence_filter)
        elif tab == "tab-sat-pro":
            return render_sat_pro(mois_filter, uc_filter, agence_filter)
        return html.Div("Sélectionnez un onglet")

    # ── Export Excel ──────────────────────────────────────
    @app.callback(
        Output("download-kpi-excel", "data"),
        Input("btn-export-kpi", "n_clicks"),
        State("kpi-tabs", "active_tab"),
        State("kpi-hist-filter", "value"),
        prevent_initial_call=True,
    )
    def export_excel(n, tab, mois_filter):
        if not n:
            return dash.no_update
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ts = datetime.now().strftime("%Y%m%d_%H%M")

        # ── Helpers styles ──────────────────────────────
        BORDER = Border(
            left=Side(style="thin",color="30363D"),right=Side(style="thin",color="30363D"),
            top=Side(style="thin",color="30363D"), bottom=Side(style="thin",color="30363D"),
        )
        def hdr(cell, txt=None):
            if txt is not None: cell.value = txt
            cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            cell.fill      = PatternFill("solid", start_color="1F497D")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = BORDER

        def uc_cell(cell, txt=None):
            if txt is not None: cell.value = txt
            cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            cell.fill      = PatternFill("solid", start_color="C0714F")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDER

        def data_cell(cell, txt=None, row_idx=0, txt_color="E6EDF3", bold=False, align="center"):
            if txt is not None: cell.value = txt
            bg = "21262D" if row_idx % 2 == 1 else "1C2128"
            cell.font      = Font(color=txt_color, bold=bold, name="Calibri", size=10)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border    = BORDER

        def auto_col_width(ws):
            for col in ws.columns:
                w = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(w + 4, 45)

        def freeze_and_zoom(ws):
            ws.freeze_panes = "A2"
            ws.sheet_view.zoomScale = 90

        # ── Étape 1 : écrire les données brutes ─────────
        buf1 = io.BytesIO()
        if tab == "tab-crc":
            df = get_crc_history()
            if mois_filter: df = df[df["mois"] == mois_filter]
            pivot = pd.DataFrame({"Message":["Aucune donnée"]}) if df.empty else (
                df.pivot_table(index=["indicateur","objectif"],columns="mois",
                               values="valeur",aggfunc="first").reset_index()
                  .rename_axis(None,axis=1)
            )
            with pd.ExcelWriter(buf1, engine="openpyxl") as w:
                pivot.to_excel(w, sheet_name="KPI CRC", index=False)
            fname = f"KPI_CRC_export_{ts}.xlsx"

        elif tab in ("tab-sat","tab-sat-pro"):
            df  = get_sat_history() if tab=="tab-sat" else get_sat_pro_history()
            lbl = "SAT" if tab=="tab-sat" else "SAT_PRO"
            if mois_filter: df = df[df["mois"] == mois_filter]
            with pd.ExcelWriter(buf1, engine="openpyxl") as w:
                if df.empty:
                    pd.DataFrame({"Message":["Aucune donnée"]}).to_excel(w, sheet_name=lbl, index=False)
                else:
                    for feuille in df["feuille"].unique():
                        dff   = df[df["feuille"]==feuille]
                        pivot = dff.pivot_table(index="entite",columns="mois",
                                                values="valeur",aggfunc="first").reset_index()
                        pivot.columns.name = None
                        # Joindre is_uc pour styling
                        if "is_uc" in dff.columns:
                            uc_info = dff.drop_duplicates("entite")[["entite","uc","is_uc"]]
                            pivot   = pivot.merge(uc_info, on="entite", how="left")
                        pivot.to_excel(w, sheet_name=feuille[:31], index=False)
            fname = f"KPIS_{lbl}_export_{ts}.xlsx"
        else:
            return dash.no_update

        # ── Étape 2 : rouvrir et appliquer les styles ───
        buf1.seek(0)
        wb = load_workbook(buf1)

        if tab == "tab-crc":
            ws = wb["KPI CRC"]
            # En-têtes
            for cell in ws[1]:
                hdr(cell)
            ws.row_dimensions[1].height = 28
            # Données
            mois_cols_idx = [j for j,c in enumerate(ws[1],1) if c.value not in ("indicateur","objectif")]
            for r_idx, row in enumerate(ws.iter_rows(min_row=2), 0):
                for cell in row:
                    col_name = ws.cell(row=1,column=cell.column).value
                    if col_name == "indicateur":
                        data_cell(cell, row_idx=r_idx, txt_color="58A6FF", bold=True, align="left")
                    elif col_name == "objectif":
                        data_cell(cell, row_idx=r_idx, txt_color="FFA657", bold=True)
                    else:
                        data_cell(cell, row_idx=r_idx)

        else:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row < 2: continue

                # Identifier colonnes
                headers = [ws.cell(row=1,column=j).value for j in range(1,ws.max_column+1)]
                entite_col = 1
                mois_start = next((j+1 for j,h in enumerate(headers) if h not in ("entite","uc","is_uc")), 2)
                uc_col     = next((j+1 for j,h in enumerate(headers) if h=="uc"),     None)
                is_uc_col  = next((j+1 for j,h in enumerate(headers) if h=="is_uc"),  None)
                # Masquer colonnes techniques uc/is_uc
                if uc_col:    ws.column_dimensions[ws.cell(1,uc_col).column_letter].hidden   = True
                if is_uc_col: ws.column_dimensions[ws.cell(1,is_uc_col).column_letter].hidden = True

                # En-têtes
                for j in range(1, ws.max_column+1):
                    h = ws.cell(row=1,column=j).value
                    if h not in ("uc","is_uc"):
                        hdr(ws.cell(row=1,column=j))
                ws.row_dimensions[1].height = 28

                # Données
                ag_idx = 0
                for r in range(2, ws.max_row+1):
                    is_uc_row = False
                    if is_uc_col:
                        v = ws.cell(row=r,column=is_uc_col).value
                        is_uc_row = str(v) in ("1","1.0","True")

                    for j in range(1, ws.max_column+1):
                        h = ws.cell(row=1,column=j).value
                        if h in ("uc","is_uc"): continue
                        cell = ws.cell(row=r,column=j)
                        if is_uc_row:
                            uc_cell(cell)
                            if j==1: cell.alignment = Alignment(horizontal="center",vertical="center")
                        else:
                            align = "left" if j==1 else "center"
                            data_cell(cell, row_idx=ag_idx, align=align)
                    if not is_uc_row:
                        ag_idx += 1

                auto_col_width(ws)
                freeze_and_zoom(ws)

        if tab == "tab-crc":
            auto_col_width(wb["KPI CRC"])
            freeze_and_zoom(wb["KPI CRC"])

        # ── Étape 3 : sauvegarder dans buffer final ──────
        buf2 = io.BytesIO()
        wb.save(buf2)
        buf2.seek(0)
        return dcc.send_bytes(buf2.getvalue(), fname)


# ── Fonctions de rendu ───────────────────────────────────────

def _gcard(fig, width=12):
    return dbc.Col(
        html.Div(dcc.Graph(figure=fig, config={"displayModeBar": False}), style=card_style),
        width=width,
    )

def empty_state(msg="Aucune donnée disponible. Importez un fichier pour commencer."):
    return html.Div([
        html.Div("📭", style={"fontSize": "52px", "textAlign": "center", "marginBottom": "12px"}),
        html.P(msg, style={"color": "#8b949e", "textAlign": "center", "fontSize": "14px"}),
    ], style={"paddingTop": "80px"})


def render_crc(mois_filter):
    df = get_crc_history()
    if df.empty:
        return empty_state("Importez le fichier KPI CRC.xlsx pour afficher les indicateurs.")
    if mois_filter:
        df = df[df["mois"] == mois_filter]

    rows = []
    # KPI cards pour le dernier mois disponible
    last = df["mois"].max()
    df_last = df[df["mois"] == last]
    kpis_importants = ["Taux de joignabilité", "Taux d'occupation", "Taux d'interaction",
                       "Qualité d'écoute", "Taux de réponse à la 1ère demande"]
    colors_cycle = [C["accent"], C["green"], C["orange"], C["purple"] if "purple" in C else "#d2a8ff", C["red"]]

    kpi_cards = []
    for i, indic in enumerate(kpis_importants):
        row = df_last[df_last["indicateur"].str.contains(indic[:15], case=False, na=False)]
        val = row.iloc[0]["valeur"] if not row.empty else "N/A"
        obj = row.iloc[0]["objectif"] if not row.empty else ""
        kpi_cards.append(
            dbc.Col(html.Div([
                html.Div(indic[:25], style={"fontSize": "10px", "color": "#8b949e",
                                             "textTransform": "uppercase", "letterSpacing": "0.7px",
                                             "fontWeight": "600", "marginBottom": "6px"}),
                html.Div(val, style={"fontSize": "24px", "fontWeight": "800",
                                     "color": colors_cycle[i % len(colors_cycle)], "lineHeight": "1"}),
                html.Div(f"Obj: {obj}" if obj else "", style={"fontSize": "11px", "color": "#8b949e", "marginTop": "4px"}),
            ], style={**card_style, "borderLeft": f"3px solid {colors_cycle[i % len(colors_cycle)]}",
                      "padding": "14px 16px", "marginBottom": "12px"}),
            width=2)
        )
    rows.append(dbc.Row(kpi_cards, className="mb-3"))

    # Tableau CRC embelli
    pivot = df.pivot_table(index=["indicateur", "objectif"], columns="mois", values="valeur", aggfunc="first").reset_index()
    pivot.columns.name = None
    mois_cols = [c for c in pivot.columns if c not in ("indicateur", "objectif")]

    header_cells = [
        html.Th("INDICATEUR", style={"backgroundColor":"#1F497D","color":"white","padding":"10px 14px",
                                      "textAlign":"left","fontWeight":"700","fontSize":"11px",
                                      "border":"1px solid #30363d","minWidth":"240px","letterSpacing":"0.6px"}),
        html.Th("OBJECTIF",   style={"backgroundColor":"#1F497D","color":"white","padding":"10px 14px",
                                      "textAlign":"center","fontWeight":"700","fontSize":"11px",
                                      "border":"1px solid #30363d","minWidth":"100px","letterSpacing":"0.6px"}),
    ]
    for m in mois_cols:
        header_cells.append(html.Th(m, style={
            "backgroundColor":"#1F497D","color":"white","padding":"10px 14px",
            "textAlign":"center","fontWeight":"700","fontSize":"11px",
            "border":"1px solid #30363d","minWidth":"110px",
        }))

    body_rows = []
    for i, row in pivot.iterrows():
        bg = "#1c2128" if i % 2 == 0 else "#21262d"
        cells = [
            html.Td(row["indicateur"], style={"backgroundColor":bg,"color":"#58a6ff",
                                               "fontWeight":"600","fontSize":"12px",
                                               "padding":"9px 14px","border":"1px solid #30363d"}),
            html.Td(row.get("objectif",""), style={"backgroundColor":bg,"color":"#ffa657",
                                                    "textAlign":"center","fontSize":"12px",
                                                    "padding":"9px 14px","border":"1px solid #30363d",
                                                    "fontWeight":"600"}),
        ]
        for m in mois_cols:
            v = row.get(m, "")
            val_str = "" if pd.isna(v) or str(v) in ("nan","None","") else str(v)
            # Colorer selon objectif vs valeur
            cells.append(html.Td(val_str, style={
                "backgroundColor":bg,"color":"#e6edf3","textAlign":"center",
                "fontSize":"12px","padding":"9px 14px","border":"1px solid #30363d",
            }))
        body_rows.append(html.Tr(cells))

    crc_table = html.Div(
        html.Table(
            [html.Thead(html.Tr(header_cells)), html.Tbody(body_rows)],
            style={"width":"100%","borderCollapse":"collapse","fontFamily":"'Segoe UI', sans-serif"},
        ),
        style={"overflowX":"auto","border":"1px solid #30363d","borderRadius":"8px"},
    )
    rows.append(dbc.Row([dbc.Col(html.Div([
        html.H6("Historique des indicateurs", style={"color":"#e6edf3","fontWeight":"700","marginBottom":"12px"}),
        crc_table,
    ], style=card_style), width=12)]))

    # Graphique évolution taux joignabilité
    df_indics = df[df["indicateur"].str.contains("joignab|occupation|écoute|interact", case=False, na=False)].copy()
    if not df_indics.empty:
        df_indics["valeur_num"] = pd.to_numeric(df_indics["valeur"].str.replace('%','',regex=False), errors='coerce')
        fig = px.line(df_indics, x="mois", y="valeur_num", color="indicateur", markers=True,
                      title="Évolution des taux clés (%)", template="plotly_dark",
                      color_discrete_sequence=PALETTE if 'PALETTE' in dir() else px.colors.qualitative.Set2)
        fig.update_traces(mode="lines+markers+text", textposition="top center")
        apply_layout(fig)
        rows.append(dbc.Row([_gcard(fig, 12)]))

    return html.Div(rows)


def _build_sat_table(dff, mois_cols, uc_filter):
    """
    Construit un tableau HTML fidèle à la capture :
    - Lignes UC : fond orange, texte blanc, centré, en gras
    - Lignes agences : fond alterné clair/sombre, valeurs centrées
    """
    has_uc  = "uc"    in dff.columns
    has_is_uc = "is_uc" in dff.columns

    # Construire la liste ordonnée des lignes dans l'ordre original (UC puis ses agences)
    if has_uc and has_is_uc and not uc_filter:
        # Trier : d'abord les UC (is_uc=1) puis leurs agences, groupées par UC
        dff_sorted = dff.copy()
        dff_sorted["_sort_uc"] = dff_sorted["uc"].fillna("ZZZ")
        dff_sorted["_sort_is_uc"] = dff_sorted["is_uc"].fillna(0)
        df_ordered = dff_sorted.sort_values(
            ["_sort_uc", "_sort_is_uc"], ascending=[True, False]
        ).drop_duplicates("entite").drop(columns=["_sort_uc","_sort_is_uc"])
    else:
        df_ordered = dff.drop_duplicates("entite")

    # Pivot mois en colonnes
    pivot = dff.pivot_table(index="entite", columns="mois", values="valeur",
                            aggfunc="first").reset_index()
    pivot.columns.name = None
    mois_disponibles = [c for c in pivot.columns if c != "entite"]

    # Colonnes header
    header_cells = [
        html.Th("AGENCES", style={
            "backgroundColor": "#1F497D", "color": "white",
            "padding": "10px 14px", "textAlign": "left",
            "fontWeight": "700", "fontSize": "12px",
            "border": "1px solid #30363d", "minWidth": "220px",
        })
    ]
    for m in mois_disponibles:
        header_cells.append(html.Th(m, style={
            "backgroundColor": "#1F497D", "color": "white",
            "padding": "10px 14px", "textAlign": "center",
            "fontWeight": "700", "fontSize": "12px",
            "border": "1px solid #30363d", "minWidth": "100px",
        }))

    # Lignes de données
    body_rows = []
    ag_idx = 0  # pour alterner fond agences

    # Récupérer les UCs et leurs agences dans l'ordre
    if has_uc and has_is_uc and not uc_filter:
        uc_set = dff[dff["is_uc"] == 1]["entite"].unique().tolist()
    else:
        uc_set = []

    seen = set()
    for _, row in df_ordered.iterrows():
        entite = row["entite"]
        if entite in seen:
            continue
        seen.add(entite)

        is_uc_row = has_is_uc and (row["is_uc"] == 1 if "is_uc" in row.index else False)

        # Valeurs pour chaque mois
        val_row = pivot[pivot["entite"] == entite]

        if is_uc_row:
            # Ligne UC — style orange comme la capture
            cells = [html.Td(entite, style={
                "backgroundColor": "#C0714F", "color": "white",
                "fontWeight": "700", "fontSize": "12px",
                "padding": "9px 14px", "textAlign": "center",
                "border": "1px solid #30363d",
            })]
            for m in mois_disponibles:
                v = val_row[m].values[0] if not val_row.empty and m in val_row.columns else ""
                cells.append(html.Td(
                    "" if pd.isna(v) or str(v) in ("nan","None","") else str(v),
                    style={
                        "backgroundColor": "#C0714F", "color": "white",
                        "fontWeight": "700", "fontSize": "12px",
                        "padding": "9px 14px", "textAlign": "center",
                        "border": "1px solid #30363d",
                    }
                ))
            body_rows.append(html.Tr(cells))
        else:
            # Ligne agence — fond alterné
            bg = "#1c2128" if ag_idx % 2 == 0 else "#21262d"
            ag_idx += 1
            cells = [html.Td(entite, style={
                "backgroundColor": bg, "color": "#e6edf3",
                "fontSize": "12px", "padding": "8px 14px",
                "border": "1px solid #30363d",
            })]
            for m in mois_disponibles:
                v = val_row[m].values[0] if not val_row.empty and m in val_row.columns else ""
                cells.append(html.Td(
                    "" if pd.isna(v) or str(v) in ("nan","None","") else str(v),
                    style={
                        "backgroundColor": bg, "color": "#e6edf3",
                        "fontSize": "12px", "padding": "8px 14px",
                        "textAlign": "center", "border": "1px solid #30363d",
                    }
                ))
            body_rows.append(html.Tr(cells))

    return html.Div(
        html.Table(
            [html.Thead(html.Tr(header_cells)), html.Tbody(body_rows)],
            style={"width": "100%", "borderCollapse": "collapse",
                   "fontFamily": "'Segoe UI', sans-serif"},
        ),
        style={"overflowX": "auto", "marginTop": "16px",
               "border": "1px solid #30363d", "borderRadius": "8px"},
    )


def render_sat(mois_filter, uc_filter=None, agence_filter=None):
    try:
        agences_filtre = None
        if agence_filter:
            agences_filtre = agence_filter if isinstance(agence_filter, list) else [agence_filter]
        elif uc_filter:
            agences_filtre = get_agences_for_uc(uc_filter)

        df = get_sat_history(uc=uc_filter, agences=agences_filtre)
        if df.empty:
            return empty_state("Importez le fichier KPIS SAT.xlsx pour afficher les indicateurs.")
        if mois_filter:
            df = df[df["mois"] == mois_filter]
        if df.empty:
            return empty_state("Aucune donnée pour cette combinaison de filtres.")

        has_uc    = "uc"    in df.columns
        has_is_uc = "is_uc" in df.columns

        feuilles = df["feuille"].unique().tolist()
        tabs = []
        for feuille in feuilles:
            dff  = df[df["feuille"] == feuille].copy()
            last = dff["mois"].max()
            mois_cols = sorted(dff["mois"].unique().tolist())

            content_tab = []
            titre = feuille + (f" — {uc_filter}" if uc_filter else "")
            content_tab.append(html.H6(titre, style={"color":"#e6edf3","fontWeight":"700","marginBottom":"12px"}))

            # Tableau fidèle à la capture
            content_tab.append(_build_sat_table(dff, mois_cols, uc_filter))

            tabs.append(dbc.Tab(html.Div(content_tab, style={"paddingTop": "16px"}),
                                label=feuille[:25], tab_id=f"sat-{feuille}"))

        if not tabs:
            return empty_state()
        return dbc.Tabs(tabs, active_tab=f"sat-{feuilles[0]}")

    except Exception as _e:
        import traceback as _tb
        return html.Div([
            html.Div("❌ Erreur affichage SAT — vérifiez les logs ci-dessous",
                     style={"color":"#f78166","fontWeight":"700","marginBottom":"8px"}),
            html.Pre(str(_e),
                     style={"color":"#ffa657","fontSize":"12px","whiteSpace":"pre-wrap"}),
            html.Pre(_tb.format_exc(),
                     style={"color":"#8b949e","fontSize":"10px","whiteSpace":"pre-wrap",
                            "maxHeight":"200px","overflowY":"auto",
                            "backgroundColor":"#0d1117","padding":"8px","borderRadius":"6px"}),
        ], style={"backgroundColor":"#1c2128","padding":"16px","borderRadius":"8px",
                  "border":"1px solid #f78166","marginTop":"16px"})


def render_sat_pro(mois_filter, uc_filter=None, agence_filter=None):
    """Même rendu que render_sat mais depuis kpi_sat_pro_v2."""
    f1 = render_feuille1("sat_pro", mois_filter)

    agences_filtre = None
    if agence_filter:
        agences_filtre = agence_filter if isinstance(agence_filter, list) else [agence_filter]
    elif uc_filter:
        agences_filtre = get_agences_for_uc(uc_filter)

    df = get_sat_pro_history(uc=uc_filter, agences=agences_filtre)
    if df.empty:
        return empty_state("Importez le fichier KPIS SAT PRO.xlsx pour afficher les indicateurs.")
    if mois_filter:
        df = df[df["mois"] == mois_filter]
    if df.empty:
        return empty_state("Aucune donnée pour cette combinaison de filtres.")

    has_uc    = "uc"    in df.columns
    has_is_uc = "is_uc" in df.columns

    feuilles = df["feuille"].unique().tolist()
    tabs = []
    for feuille in feuilles:
        dff  = df[df["feuille"] == feuille].copy()
        last = dff["mois"].max()
        mois_cols = sorted(dff["mois"].unique().tolist())

        content = []
        titre = feuille + (f" — {uc_filter}" if uc_filter else "")
        content.append(html.H6(titre, style={"color":"#e6edf3","fontWeight":"700","marginBottom":"12px"}))

        content.append(_build_sat_table(dff, mois_cols, uc_filter))
        tabs.append(dbc.Tab(html.Div(content, style={"paddingTop": "16px"}),
                            label=feuille[:25], tab_id=f"satpro-{feuille}"))

    if not tabs:
        result = [f1] if f1 else []
        return html.Div(result) if result else empty_state()
    tabs_widget = dbc.Tabs(tabs, active_tab=f"satpro-{feuilles[0]}")
    return html.Div([f1, tabs_widget] if f1 else [tabs_widget])


# ── LAYOUT RACINE ─────────────────────────────────────────────────────────────
# On utilise display:none / display:block — PAS de routing, PAS de dcc.Location
# Les 3 pages sont toujours dans le DOM, on les affiche/masque simplement
app.layout = html.Div([
    html.Div(accueil,  id="page-accueil",  style={"display":"block"}),
    html.Div(dashboard, id="page-dashboard", style={"display":"none"}),
    html.Div(module,   id="page-module",   style={"display":"none"}),
])

# ── CALLBACK NAVIGATION ───────────────────────────────────────────────────────
@app.callback(
    Output("page-accueil",  "style"),
    Output("page-dashboard", "style"),
    Output("page-module",   "style"),
    Input("btn-open-dashboard",      "n_clicks"),
    Input("btn-open-module",         "n_clicks"),
    Input("btn-retour-accueil",      "n_clicks"),
    Input("btn-retour-accueil-module","n_clicks"),
    prevent_initial_call=True,
)
def navigate(n1, n2, n3, n4):
    show   = {"display": "block"}
    hide   = {"display": "none"}
    tid = ctx.triggered_id
    if tid == "btn-open-dashboard":
        return hide, show, hide
    if tid == "btn-open-module":
        return hide, hide, show
    # retour accueil (depuis dashboard ou module)
    return show, hide, hide

# ── CALLBACKS DASHBOARD (identiques à l'original) ─────────────────────────────
import traceback

@app.callback(
    Output('stored-data', 'data'),
    Output('upload-status', 'children'),
    Output('debug-panel', 'children'),
    Input('upload-data', 'contents'),
    State('upload-data', 'filename'),
    prevent_initial_call=True
)
def store_uploaded_data(contents, filename):
    ts = datetime.now().strftime("%H:%M:%S")
    logs = []

    if not contents or not filename:
        return dash.no_update, "", f"[{ts}] Aucun fichier recu."

    logs.append(f"[{ts}] Fichier recu : {filename}")
    logs.append(f"[{ts}] Taille base64 : {len(contents)} caracteres")

    try:
        df_raw, error = parse_contents(contents, filename)
        if error:
            logs.append(f"[{ts}] ERREUR parse : {error}")
            status = html.Div(f"Erreur : {error}", style={"color":"#f78166","fontSize":"12px","marginTop":"8px"})
            return dash.no_update, status, "\n".join(logs)
        logs.append(f"[{ts}] Parse OK - {len(df_raw)} lignes, {len(df_raw.columns)} colonnes")
        logs.append(f"[{ts}] Colonnes: {list(df_raw.columns[:6])}")
    except Exception as e:
        tb = traceback.format_exc()
        logs.append(f"[{ts}] EXCEPTION parse: {str(e)}")
        logs.append(tb)
        status = html.Div(f"Erreur lecture: {str(e)}", style={"color":"#f78166","fontSize":"12px","marginTop":"8px"})
        return dash.no_update, status, "\n".join(logs)

    try:
        df = load_data(df_raw)
        logs.append(f"[{ts}] load_data OK - {len(df)} lignes finales")
        if 'NATURE' in df.columns:
            logs.append(f"[{ts}] NATURE: {df['NATURE'].unique().tolist()}")
        if 'DELAI_RECLAMATION' in df.columns:
            logs.append(f"[{ts}] DELAI: {df['DELAI_RECLAMATION'].value_counts().to_dict()}")
        status = html.Div(
            f"Fichier charge : {filename} ({len(df)} lignes)",
            style={"color":"#3fb950","fontSize":"12px","marginTop":"8px","fontWeight":"600"}
        )
        return df.to_json(date_format='iso', orient='split'), status, "\n".join(logs)
    except Exception as e:
        tb = traceback.format_exc()
        logs.append(f"[{ts}] EXCEPTION load_data: {str(e)}")
        logs.append(tb)
        status = html.Div(f"Erreur traitement: {str(e)}", style={"color":"#f78166","fontSize":"12px","marginTop":"8px"})
        return dash.no_update, status, "\n".join(logs)


@app.callback(
    Output("debug-panel", "children", allow_duplicate=True),
    Input("btn-test-cb", "n_clicks"),
    prevent_initial_call=True,
)
def test_callback(n):
    ts = datetime.now().strftime("%H:%M:%S")
    import sys, platform
    lines = [
        f"[{ts}] Callback OK - les callbacks fonctionnent",
        f"[{ts}] Python : {sys.version.split()[0]}",
        f"[{ts}] Platform : {platform.system()}",
        f"[{ts}] Pandas : {pd.__version__}",
    ]
    try:
        import openpyxl
        lines.append(f"[{ts}] openpyxl : {openpyxl.__version__} - OK")
    except ImportError:
        lines.append(f"[{ts}] openpyxl : MANQUANT - les .xlsx ne fonctionneront pas !")
    try:
        import unidecode
        lines.append(f"[{ts}] unidecode : OK")
    except ImportError:
        lines.append(f"[{ts}] unidecode : MANQUANT !")
    return "\n".join(lines)


register_kpi_callbacks(app)

if __name__ == '__main__':
    app.run(debug=True)
