# -*- coding: utf-8 -*-
"""
app.py
======
Application principale - Gestion Factures CIE V3.0
"""

import streamlit as st
import pandas as pd
import io
from datetime import datetime
import plotly.graph_objects as go

# Imports des modules
from models import load_central, save_central, COLONNES_BASE_CENTRALE
from import_bt import page_import_bt
from import_ht import page_import_ht
from generation import page_generation_fichiers
from non_enregistrees import page_non_enregistrees

# 🔐 Module d'authentification
import auth

st.set_page_config(
    page_title="Gestion Factures CIE - V3",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# AUTHENTIFICATION
# ============================================

# Initialiser le système d'authentification
auth.init_users_file()

# Vérifier si l'utilisateur est connecté
if not auth.verifier_session():
    # Afficher la page de connexion
    auth.page_connexion()
    st.stop()  # Arrêter l'exécution si non connecté

# ============================================
# APPLICATION (accessible uniquement si connecté)
# ============================================

# CSS personnalisé
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .main-title {
        color: white;
        font-size: 2.5rem;
        font-weight: bold;
        margin: 0;
        text-align: center;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        text-align: center;
    }
</style>
""", unsafe_allow_html=True)

# Initialisation
if 'df_central' not in st.session_state:
    st.session_state.df_central = load_central()
    st.session_state.base_rechargee = True  # Marquer le premier chargement
else:
    st.session_state.base_rechargee = False

# Header
st.markdown("""
<div class="main-header">
    <h1 class="main-title" style="color: #f0f0f0">📊 Gestion Factures CIE</h1>
    <p class="main-subtitle" style="color: #f0f0f0; text-align: center; margin-top: 0.5rem;">
        Version 3.0 - Gestion simplifiée avec pièces comptables
    </p>
</div>
""", unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.markdown("### 📋 Navigation")
    
    # Menu de base
    menu_items = [
        "📊 Base Centrale", 
        "📋 Non Enregistrées", 
        "🔄 Import Factures BT", 
        "🔄 Import Factures HT", 
        "📈 Statistiques", 
        "⚙️ Génération Fichiers"
    ]
    
    # Ajouter gestion utilisateurs si admin
    if auth.est_admin():
        menu_items.append("👥 Gestion Utilisateurs")
    
    page = st.radio(
        "Menu principal",
        menu_items
    )
    
    st.markdown("---")
    st.markdown("### 📊 Informations")
    
    df_central = st.session_state.df_central
    nb_lignes = len(df_central[df_central['DATE'].notna() & (df_central['DATE'] != '')])
    st.metric("📝 Lignes", nb_lignes)
    
    if 'DATE' in df_central.columns:
        periodes = df_central['DATE'].dropna().nunique()
        st.metric("📅 Périodes", periodes)
    
    # Afficher info utilisateur et bouton déconnexion
    auth.afficher_user_info()
    
    st.markdown("---")
    st.markdown("### 💾 Sauvegarde")
    st.info("✅ **Backup automatique**\nChaque modification est sauvegardée en double :\n- `data_centrale.pkl`\n- `Base_Centrale.xlsx`")
    
    with st.expander("🛟 En cas de problème"):
        st.markdown("""
        **Si erreur "Erreur pickle" :**
        
        1. Ne pas paniquer ! ✅
        2. Supprimer `data_centrale.pkl`
        3. Relancer l'application
        4. ✅ **Vos données seront rechargées depuis `Base_Centrale.xlsx`**
        
        **Protection :**
        - Backup Excel créé à chaque sauvegarde
        - Historique complet préservé
        - Aucune perte de données possible
        """)


# ===== PAGES =====

if page == "📊 Base Centrale":
    st.markdown("## 📊 Base Centrale")
    st.markdown("---")
    
    st.info("""
    📌 **Colonnes de la base centrale** :
    - UC, CODE RED, CODE AGCE, SITES, IDENTIFIANT
    - TENSION, DATE, CONSO, MONTANT, DATE_COMPLEMENTAIRE
    
    💡 Vous pouvez ajouter, modifier ou supprimer des lignes manuellement.
    """)
    
    df_central = st.session_state.df_central
    
    # Statistiques
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    
    with col_stat1:
        nb_total = len(df_central[df_central['DATE'].notna() & (df_central['DATE'] != '')])
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: #667eea;">📝</h3>
            <h2>{nb_total}</h2>
            <p style="color: #666;">Lignes totales</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col_stat2:
        sites_uniques = df_central['IDENTIFIANT'].nunique() if 'IDENTIFIANT' in df_central.columns else 0
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: #667eea;">🏢</h3>
            <h2>{sites_uniques}</h2>
            <p style="color: #666;">Sites uniques</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col_stat3:
        periodes = df_central['DATE'].dropna().nunique() if 'DATE' in df_central.columns else 0
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: #667eea;">📅</h3>
            <h2>{periodes}</h2>
            <p style="color: #666;">Périodes</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col_stat4:
        total = df_central['MONTANT'].sum() if 'MONTANT' in df_central.columns else 0
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: #667eea;">💰</h3>
            <h2>{total/1000:.0f}K</h2>
            <p style="color: #666;">Total FCFA</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Filtres
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    
    with col_f1:
        if 'UC' in df_central.columns:
            ucs = ['Tous'] + sorted(df_central['UC'].dropna().unique().tolist())
            uc_filter = st.selectbox("Filtrer par UC", ucs)
        else:
            uc_filter = 'Tous'
    
    with col_f2:
        if 'DATE' in df_central.columns:
            dates = ['Tous'] + sorted(df_central['DATE'].dropna().unique().tolist(), reverse=True)
            date_filter = st.selectbox("Filtrer par DATE", dates)
        else:
            date_filter = 'Tous'
    
    with col_f3:
        if 'TENSION' in df_central.columns:
            tensions = ['Tous'] + sorted(df_central['TENSION'].dropna().unique().tolist())
            tension_filter = st.selectbox("Filtrer par TENSION", tensions)
        else:
            tension_filter = 'Tous'
    
    with col_f4:
        if 'STATUT' in df_central.columns:
            statuts = ['Tous'] + sorted(df_central['STATUT'].dropna().unique().tolist())
            statut_filter = st.selectbox("Filtrer par STATUT", statuts)
        else:
            statut_filter = 'Tous'
    
    # Appliquer filtres
    df_filtered = df_central.copy()
    
    if uc_filter != 'Tous':
        df_filtered = df_filtered[df_filtered['UC'] == uc_filter]
    if date_filter != 'Tous':
        df_filtered = df_filtered[df_filtered['DATE'] == date_filter]
    if tension_filter != 'Tous':
        df_filtered = df_filtered[df_filtered['TENSION'] == tension_filter]
    if statut_filter != 'Tous':
        df_filtered = df_filtered[df_filtered['STATUT'] == statut_filter]
    
    st.markdown(f"### 📋 Données ({len(df_filtered)} ligne(s))")
    
    # Éditeur
    edited_df = st.data_editor(
        df_filtered[COLONNES_BASE_CENTRALE],
        use_container_width=True,
        num_rows="dynamic",
        height=500,
        key="editor_central"
    )
    
    # Actions
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("💾 Sauvegarder modifications", type="primary", use_container_width=True):
            try:
                # Mettre à jour les lignes modifiées
                for idx in edited_df.index:
                    if idx in st.session_state.df_central.index:
                        st.session_state.df_central.loc[idx, COLONNES_BASE_CENTRALE] = edited_df.loc[idx, COLONNES_BASE_CENTRALE]
                
                # Ajouter les nouvelles lignes
                nouvelles_lignes = edited_df[~edited_df.index.isin(st.session_state.df_central.index)]
                if len(nouvelles_lignes) > 0:
                    st.session_state.df_central = pd.concat([st.session_state.df_central, nouvelles_lignes], ignore_index=True)
                    st.success(f"➕ {len(nouvelles_lignes)} ligne(s) ajoutée(s)")
                
                save_central(st.session_state.df_central)
                st.success("✅ Base centrale sauvegardée !")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Erreur lors de la sauvegarde : {str(e)}")
                st.warning("⚠️ Les modifications sont conservées en mémoire mais non sauvegardées sur disque.")
                st.info("💡 Essayez d'exporter en Excel pour ne pas perdre vos données.")
    
    with col2:
        # Gestion du statut (désactivation de sites)
        if len(df_filtered) > 0:
            with st.popover("🔄 Gérer statut sites", use_container_width=True):
                st.markdown("### Activer / Désactiver des sites")
                st.info("💡 Les sites **INACTIF** sont conservés dans l'historique mais exclus des pièces comptables.")
                
                # Option 1: Par IDENTIFIANT
                if 'IDENTIFIANT' in df_filtered.columns:
                    st.markdown("#### Par IDENTIFIANT")
                    identifiants = df_filtered['IDENTIFIANT'].unique().tolist()
                    ids_a_modifier = st.multiselect(
                        "Sélectionner les IDENTIFIANT",
                        identifiants,
                        key="ids_statut"
                    )
                    
                    if ids_a_modifier:
                        nouveau_statut = st.radio(
                            "Nouveau statut",
                            ['ACTIF', 'INACTIF'],
                            key="nouveau_statut_id"
                        )
                        
                        if st.button(f"✏️ Mettre à {nouveau_statut}", type="secondary"):
                            indices = df_filtered[df_filtered['IDENTIFIANT'].isin(ids_a_modifier)].index
                            st.session_state.df_central.loc[indices, 'STATUT'] = nouveau_statut
                            save_central(st.session_state.df_central)
                            st.success(f"✅ {len(indices)} ligne(s) → {nouveau_statut}")
                            st.rerun()
                
                st.markdown("---")
                
                # Option 2: Par DATE
                if 'DATE' in df_filtered.columns:
                    st.markdown("#### Par DATE")
                    dates = df_filtered['DATE'].dropna().unique().tolist()
                    if dates:
                        dates_a_modifier = st.multiselect(
                            "Sélectionner les DATE",
                            dates,
                            key="dates_statut"
                        )
                        
                        if dates_a_modifier:
                            nouveau_statut_date = st.radio(
                                "Nouveau statut",
                                ['ACTIF', 'INACTIF'],
                                key="nouveau_statut_date"
                            )
                            
                            if st.button(f"✏️ Mettre à {nouveau_statut_date}", type="secondary", key="btn_statut_date"):
                                indices = df_filtered[df_filtered['DATE'].isin(dates_a_modifier)].index
                                st.session_state.df_central.loc[indices, 'STATUT'] = nouveau_statut_date
                                save_central(st.session_state.df_central)
                                st.success(f"✅ {len(indices)} ligne(s) → {nouveau_statut_date}")
                                st.rerun()
    
    with col3:
        # Export Excel professionnel
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Export données
            df_export = df_filtered[COLONNES_BASE_CENTRALE].fillna('')
            df_export.to_excel(writer, sheet_name='Base_Centrale', index=False, startrow=5)
            
            workbook = writer.book
            worksheet = workbook['Base_Centrale']
            
            # Styles
            header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
            header_font = Font(bold=True, size=14, name='Calibri', color='FFFFFF')
            col_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            col_header_font = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
            data_font = Font(size=10, name='Calibri')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # TITRE
            worksheet.merge_cells('A1:M1')
            cell_titre = worksheet['A1']
            cell_titre.value = "BASE CENTRALE - GESTION FACTURES CIE"
            cell_titre.font = header_font
            cell_titre.fill = header_fill
            cell_titre.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.row_dimensions[1].height = 30
            
            # INFO
            current_date = datetime.now().strftime('%d/%m/%Y %H:%M')
            worksheet['A3'] = "Date d'export:"
            worksheet['B3'] = current_date
            worksheet['A3'].font = Font(bold=True, size=10)
            worksheet['B3'].font = data_font
            
            worksheet['D3'] = "Nombre de lignes:"
            worksheet['E3'] = len(df_export)
            worksheet['D3'].font = Font(bold=True, size=10)
            worksheet['E3'].font = data_font
            
            worksheet['G3'] = "Filtre période:"
            worksheet['H3'] = filtre_periode if filtre_periode != 'Toutes' else 'Aucun'
            worksheet['G3'].font = Font(bold=True, size=10)
            worksheet['H3'].font = data_font
            
            # Style en-têtes colonnes (ligne 6 car startrow=5)
            for col_idx, col_name in enumerate(df_export.columns, start=1):
                cell = worksheet.cell(row=6, column=col_idx)
                cell.fill = col_header_fill
                cell.font = col_header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            worksheet.row_dimensions[6].height = 35
            
            # Style données
            for row_idx in range(7, 7 + len(df_export)):
                for col_idx in range(1, len(df_export.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    
                    # Alignement spécifique
                    col_name = df_export.columns[col_idx - 1]
                    if col_name in ['MONTANT', 'CONSO', 'PSABON', 'PSATTEINTE']:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if cell.value and col_name == 'MONTANT':
                            try:
                                cell.number_format = '#,##0'
                            except:
                                pass
            
            # Largeurs colonnes
            col_widths = {
                'UC': 12, 'CODE RED': 12, 'CODE AGCE': 12,
                'SITES': 30, 'IDENTIFIANT': 18, 'TENSION': 10,
                'DATE': 12, 'CONSO': 12, 'MONTANT': 15,
                'DATE_COMPLEMENTAIRE': 18, 'STATUT': 12,
                'PSABON': 12, 'PSATTEINTE': 12
            }
            
            for col_idx, col_name in enumerate(df_export.columns, start=1):
                col_letter = get_column_letter(col_idx)
                width = col_widths.get(col_name, 15)
                worksheet.column_dimensions[col_letter].width = width
            
            # Footer
            derniere_ligne = 7 + len(df_export) + 2
            worksheet.merge_cells(f'A{derniere_ligne}:M{derniere_ligne}')
            cell_footer = worksheet[f'A{derniere_ligne}']
            cell_footer.value = f"Document généré le {current_date} - Société Générale Côte d'Ivoire"
            cell_footer.font = Font(size=9, italic=True, color='666666')
            cell_footer.alignment = Alignment(horizontal='center')
        
        output.seek(0)
        
        st.download_button(
            "📥 Exporter Excel",
            data=output,
            file_name=f"Base_Centrale_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col4:
        if st.button("🔄 Actualiser", use_container_width=True):
            st.rerun()
    
    # Afficher répartition par statut
    if 'STATUT' in df_central.columns:
        st.markdown("---")
        st.markdown("### 📊 Répartition par statut")
        
        statuts = df_central['STATUT'].value_counts()
        col_s1, col_s2, col_s3 = st.columns(3)
        
        with col_s1:
            actifs = statuts.get('ACTIF', 0)
            st.metric("✅ Sites ACTIF", actifs)
        with col_s2:
            inactifs = statuts.get('INACTIF', 0)
            st.metric("⏸️ Sites INACTIF", inactifs)
        with col_s3:
            total = len(df_central['IDENTIFIANT'].unique())
            st.metric("🏢 Total sites uniques", total)

elif page == "📋 Non Enregistrées":
    page_non_enregistrees()

elif page == "🔄 Import Factures BT":
    page_import_bt()

elif page == "🔄 Import Factures HT":
    page_import_ht()

elif page == "📈 Statistiques":
    st.markdown("## 📈 Statistiques et Évolution")
    st.markdown("---")
    
    df_central = st.session_state.df_central
    
    if 'DATE' not in df_central.columns or df_central['DATE'].isna().all():
        st.warning("⚠️ Aucune période enregistrée. Importez d'abord des factures.")
    else:
        periodes_brutes = sorted(df_central['DATE'].dropna().unique().tolist())
        
        if len(periodes_brutes) == 0:
            st.warning("⚠️ Aucune donnée disponible.")
        else:
            st.success(f"✅ {len(periodes_brutes)} période(s) disponible(s)")
            
            # Filtres
            col_f1, col_f2 = st.columns(2)
            
            with col_f1:
                if 'SITES' in df_central.columns:
                    sites = ['Tous'] + sorted(df_central['SITES'].dropna().unique().tolist())
                    site_filter = st.selectbox("🏢 Filtrer par SITE", sites)
                else:
                    site_filter = 'Tous'
            
            with col_f2:
                type_graphique = st.selectbox(
                    "⚡ Type d'analyse",
                    ["📊 Global (BT + HT)", "🔌 Basse Tension uniquement", "⚡ Haute Tension uniquement"]
                )
            
            # Appliquer les filtres
            df_filtered = df_central.copy()
            
            if site_filter != 'Tous':
                df_filtered = df_filtered[df_filtered['SITES'] == site_filter]
            
            if "Basse Tension" in type_graphique:
                df_filtered = df_filtered[df_filtered['TENSION'] == 'BASSE']
            elif "Haute Tension" in type_graphique:
                df_filtered = df_filtered[df_filtered['TENSION'] == 'HAUTE']
            
            # Grouper par DATE
            df_grouped = df_filtered.groupby('DATE').agg({
                'MONTANT': 'sum',
                'CONSO': 'sum'
            }).reset_index()
            
            df_grouped = df_grouped.sort_values('DATE')
            
            st.markdown("---")
            
            # GRAPHIQUE MONTANTS
            st.markdown("### 💰 Évolution des Montants")
            
            fig_montant = go.Figure()
            
            fig_montant.add_trace(go.Scatter(
                x=df_grouped['DATE'],
                y=df_grouped['MONTANT'],
                mode='lines+markers',
                name='Montant',
                line=dict(color='#667eea', width=3),
                marker=dict(size=10, color='#667eea'),
                hovertemplate='<b>%{x}</b><br>Montant: %{y:,.0f} FCFA<extra></extra>'
            ))
            
            fig_montant.update_layout(
                hovermode='x unified',
                template='plotly_white',
                height=400
            )
            fig_montant.update_xaxes(title_text="Période")
            fig_montant.update_yaxes(title_text="Montant (FCFA)")
            
            st.plotly_chart(fig_montant, use_container_width=True)
            
            # GRAPHIQUE CONSOMMATIONS + CO2
            st.markdown("### ⚡ Évolution des Consommations et Émissions CO2")
            
            # Configuration du facteur d'émission
            col_factor, col_info = st.columns([1, 3])
            
            with col_factor:
                facteur_emission = st.number_input(
                    "Facteur CO2 (kg/kWh)",
                    min_value=0.0,
                    max_value=2.0,
                    value=0.5,
                    step=0.01,
                    help="Facteur d'émission pour la Côte d'Ivoire"
                )
            
            with col_info:
                st.info(f"💡 Émissions CO2 = Consommation × {facteur_emission} kg CO2/kWh")
            
            # Calcul des émissions CO2
            df_grouped['CO2'] = df_grouped['CONSO'] * facteur_emission
            
            # Créer le graphique avec deux axes Y
            fig_conso = go.Figure()
            
            # Consommation (axe Y principal)
            fig_conso.add_trace(go.Scatter(
                x=df_grouped['DATE'],
                y=df_grouped['CONSO'],
                mode='lines+markers',
                name='Consommation (kWh)',
                line=dict(color='#f5576c', width=3),
                marker=dict(size=10, color='#f5576c'),
                yaxis='y',
                hovertemplate='<b>%{x}</b><br>Conso: %{y:,.0f} kWh<extra></extra>'
            ))
            
            # Émissions CO2 (axe Y secondaire)
            fig_conso.add_trace(go.Scatter(
                x=df_grouped['DATE'],
                y=df_grouped['CO2'],
                mode='lines+markers',
                name='Émissions CO2 (kg)',
                line=dict(color='#28a745', width=3, dash='dot'),
                marker=dict(size=10, color='#28a745', symbol='diamond'),
                yaxis='y2',
                hovertemplate='<b>%{x}</b><br>CO2: %{y:,.0f} kg<extra></extra>'
            ))
            
            fig_conso.update_layout(
                yaxis=dict(
                    title="Consommation (kWh)",
                    titlefont=dict(color='#f5576c'),
                    tickfont=dict(color='#f5576c')
                ),
                yaxis2=dict(
                    title="Émissions CO2 (kg)",
                    titlefont=dict(color='#28a745'),
                    tickfont=dict(color='#28a745'),
                    overlaying='y',
                    side='right'
                ),
                hovermode='x unified',
                template='plotly_white',
                height=450,
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1
                )
            )
            fig_conso.update_xaxes(title_text="Période")
            
            st.plotly_chart(fig_conso, use_container_width=True)
            
            # Statistiques CO2 compactes
            col_co2_1, col_co2_2, col_co2_3, col_co2_4 = st.columns(4)
            
            total_co2 = df_grouped['CO2'].sum()
            
            with col_co2_1:
                st.metric("🌍 Total CO2", f"{total_co2:,.0f} kg", help="Total des émissions")
            
            with col_co2_2:
                if total_co2 >= 1000:
                    st.metric("CO2 (tonnes)", f"{total_co2/1000:,.2f} t")
                else:
                    st.metric("Moyenne CO2", f"{df_grouped['CO2'].mean():,.0f} kg")
            
            with col_co2_3:
                arbres = total_co2 / 25
                st.metric("🌳 Arbres équiv.", f"{arbres:,.0f}", help="Arbres nécessaires pour absorber ce CO2 en 1 an")
            
            with col_co2_4:
                km_voiture = total_co2 / 0.12
                st.metric("🚗 km voiture", f"{km_voiture:,.0f}", help="Kilomètres en voiture thermique équivalents")
            
            # ============================================
            # PUISSANCES HT (si Haute Tension)
            # ============================================
            if "Haute Tension" in type_graphique or "Global" in type_graphique:
                st.markdown("---")
                st.markdown("### ⚡ Évolution des Puissances (HT uniquement)")
                
                # Filtrer uniquement HT
                df_ht_puiss = df_central[df_central['TENSION'] == 'HAUTE'].copy()
                
                if site_filter != 'Tous':
                    df_ht_puiss = df_ht_puiss[df_ht_puiss['SITES'] == site_filter]
                
                # Vérifier si colonnes puissances existent
                if 'PSABON' in df_ht_puiss.columns and 'PSATTEINTE' in df_ht_puiss.columns:
                    # Convertir en numérique
                    df_ht_puiss['PSABON'] = pd.to_numeric(df_ht_puiss['PSABON'], errors='coerce').fillna(0)
                    df_ht_puiss['PSATTEINTE'] = pd.to_numeric(df_ht_puiss['PSATTEINTE'], errors='coerce').fillna(0)
                    
                    # Filtrer les lignes avec puissances > 0
                    df_ht_puiss = df_ht_puiss[(df_ht_puiss['PSABON'] > 0) | (df_ht_puiss['PSATTEINTE'] > 0)]
                    
                    if len(df_ht_puiss) > 0:
                        if site_filter == 'Tous':
                            # Vue globale : grouper par DATE
                            df_puiss_grouped = df_ht_puiss.groupby('DATE').agg({
                                'PSABON': 'mean',
                                'PSATTEINTE': 'mean'
                            }).reset_index()
                            
                            df_puiss_grouped = df_puiss_grouped.sort_values('DATE')
                            
                            fig_puiss = go.Figure()
                            
                            fig_puiss.add_trace(go.Scatter(
                                x=df_puiss_grouped['DATE'],
                                y=df_puiss_grouped['PSABON'],
                                mode='lines+markers',
                                name='Puissance Souscrite',
                                line=dict(color='#667eea', width=3),
                                marker=dict(size=10),
                                hovertemplate='<b>%{x}</b><br>PS Souscrite: %{y:,.1f} kVA<extra></extra>'
                            ))
                            
                            fig_puiss.add_trace(go.Scatter(
                                x=df_puiss_grouped['DATE'],
                                y=df_puiss_grouped['PSATTEINTE'],
                                mode='lines+markers',
                                name='Puissance Atteinte',
                                line=dict(color='#f5576c', width=3),
                                marker=dict(size=10),
                                hovertemplate='<b>%{x}</b><br>PS Atteinte: %{y:,.1f} kVA<extra></extra>'
                            ))
                            
                            fig_puiss.update_layout(
                                hovermode='x unified',
                                template='plotly_white',
                                height=400,
                                legend=dict(
                                    orientation="h",
                                    yanchor="bottom",
                                    y=1.02,
                                    xanchor="right",
                                    x=1
                                )
                            )
                            fig_puiss.update_xaxes(title_text="Période")
                            fig_puiss.update_yaxes(title_text="Puissance (kVA)")
                            
                            st.plotly_chart(fig_puiss, use_container_width=True)
                            
                            # Tableau récapitulatif
                            with st.expander("📊 Détails par période"):
                                st.dataframe(
                                    df_puiss_grouped.rename(columns={
                                        'DATE': 'Période',
                                        'PSABON': 'PS Souscrite (kVA)',
                                        'PSATTEINTE': 'PS Atteinte (kVA)'
                                    }),
                                    use_container_width=True,
                                    hide_index=True
                                )
                        
                        else:
                            # Vue par site : grouper par DATE pour ce site
                            df_site_puiss = df_ht_puiss.groupby('DATE').agg({
                                'PSABON': 'first',
                                'PSATTEINTE': 'first'
                            }).reset_index()
                            
                            df_site_puiss = df_site_puiss.sort_values('DATE')
                            
                            fig_puiss_site = go.Figure()
                            
                            fig_puiss_site.add_trace(go.Scatter(
                                x=df_site_puiss['DATE'],
                                y=df_site_puiss['PSABON'],
                                mode='lines+markers',
                                name='Puissance Souscrite',
                                line=dict(color='#667eea', width=3),
                                marker=dict(size=12),
                                hovertemplate='<b>%{x}</b><br>PS Souscrite: %{y:,.1f} kVA<extra></extra>'
                            ))
                            
                            fig_puiss_site.add_trace(go.Scatter(
                                x=df_site_puiss['DATE'],
                                y=df_site_puiss['PSATTEINTE'],
                                mode='lines+markers',
                                name='Puissance Atteinte',
                                line=dict(color='#f5576c', width=3),
                                marker=dict(size=12),
                                hovertemplate='<b>%{x}</b><br>PS Atteinte: %{y:,.1f} kVA<extra></extra>'
                            ))
                            
                            fig_puiss_site.update_layout(
                                title=f"Site : {site_filter}",
                                hovermode='x unified',
                                template='plotly_white',
                                height=400,
                                legend=dict(
                                    orientation="h",
                                    yanchor="bottom",
                                    y=1.02,
                                    xanchor="right",
                                    x=1
                                )
                            )
                            fig_puiss_site.update_xaxes(title_text="Période")
                            fig_puiss_site.update_yaxes(title_text="Puissance (kVA)")
                            
                            st.plotly_chart(fig_puiss_site, use_container_width=True)
                            
                            # Calcul taux utilisation
                            df_site_puiss['Taux_Utilisation'] = (df_site_puiss['PSATTEINTE'] / df_site_puiss['PSABON'] * 100).round(1)
                            
                            # Alertes
                            derniere_ligne = df_site_puiss.iloc[-1]
                            taux = derniere_ligne['Taux_Utilisation']
                            
                            col_a1, col_a2, col_a3 = st.columns(3)
                            
                            with col_a1:
                                st.metric("PS Souscrite", f"{derniere_ligne['PSABON']:.1f} kVA")
                            
                            with col_a2:
                                st.metric("PS Atteinte", f"{derniere_ligne['PSATTEINTE']:.1f} kVA")
                            
                            with col_a3:
                                if taux > 100:
                                    st.metric("Taux Utilisation", f"{taux:.1f}%", delta="⚠️ Dépassement")
                                elif taux > 90:
                                    st.metric("Taux Utilisation", f"{taux:.1f}%", delta="🟠 Proche limite")
                                else:
                                    st.metric("Taux Utilisation", f"{taux:.1f}%", delta="✅ Normal")
                            
                            # Tableau détaillé
                            with st.expander("📊 Historique détaillé"):
                                st.dataframe(
                                    df_site_puiss.rename(columns={
                                        'DATE': 'Période',
                                        'PSABON': 'PS Souscrite (kVA)',
                                        'PSATTEINTE': 'PS Atteinte (kVA)',
                                        'Taux_Utilisation': 'Taux Utilisation (%)'
                                    }),
                                    use_container_width=True,
                                    hide_index=True
                                )
                    else:
                        st.info("💡 Aucune donnée de puissance disponible pour ce filtre.")
                else:
                    st.info("💡 Les colonnes PSABON et PSATTEINTE ne sont pas présentes. Assurez-vous d'importer des factures HT avec ces données.")

elif page == "⚙️ Génération Fichiers":
    page_generation_fichiers()

elif page == "👥 Gestion Utilisateurs":
    auth.page_gestion_users()

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; padding: 2rem; color: #666;'>
    <p><strong>Système de traitement des factures CIE</strong> - Version 3.0</p>
    <p style='font-size: 0.8rem;'>🔒 Connexion sécurisée activée</p>
</div>
""", unsafe_allow_html=True)
