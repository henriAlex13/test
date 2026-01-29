"""
generation.py
=============
Génération des pièces comptables avec export Excel professionnel
"""

import pandas as pd
import streamlit as st
import io
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models import generer_piece_comptable


def export_piece_comptable_excel(df, type_tension="BT"):
    """
    Exporte la pièce comptable au format Excel professionnel
    
    Args:
        df: DataFrame pièce comptable
        type_tension: "BT" ou "HT"
    
    Returns:
        BytesIO avec le fichier Excel
    """
    output = io.BytesIO()
    
    df = df.fillna('')
    
    # Couleurs selon le type
    if type_tension == "BT":
        header_color = '2F5597'
        col_header_color = '4472C4'
        subtitle_color = '2F5597'
    else:
        header_color = 'C65911'
        col_header_color = 'ED7D31'
        subtitle_color = 'C65911'
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet('Piece_Comptable')
        
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        
        # Styles
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        header_font = Font(bold=True, size=16, color='FFFFFF', name='Calibri')
        header_align = Alignment(horizontal='center', vertical='center')
        
        subtitle_font = Font(bold=True, size=12, color=subtitle_color, name='Calibri')
        subtitle_align = Alignment(horizontal='center', vertical='center')
        
        info_font = Font(bold=True, size=10, name='Calibri')
        info_align = Alignment(horizontal='left', vertical='center')
        
        col_header_fill = PatternFill(start_color=col_header_color, end_color=col_header_color, fill_type='solid')
        col_header_font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
        col_header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        data_font = Font(size=10, name='Calibri')
        
        nb_cols = len(df.columns)
        
        # EN-TÊTE
        worksheet.merge_cells(f'A1:{get_column_letter(nb_cols)}2')
        cell = worksheet['A1']
        cell.value = "⚡ COMPAGNIE IVOIRIENNE D'ÉLECTRICITÉ (CIE)"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        worksheet.row_dimensions[1].height = 25
        worksheet.row_dimensions[2].height = 25
        
        # SOUS-TITRE
        worksheet.merge_cells(f'A3:{get_column_letter(nb_cols)}3')
        cell = worksheet['A3']
        tension_txt = "BASSE TENSION (BT)" if type_tension == "BT" else "HAUTE TENSION (HT)"
        cell.value = f"PIÈCE COMPTABLE {tension_txt}"
        cell.font = subtitle_font
        cell.alignment = subtitle_align
        worksheet.row_dimensions[3].height = 20
        
        # Date et Montant
        current_date = datetime.now().strftime('%d/%m/%Y %H:%M')
        montant_total = df['MONTANT'].sum() if 'MONTANT' in df.columns else 0
        
        worksheet['A4'] = "Date d'édition:"
        worksheet['B4'] = current_date
        worksheet['D4'] = "Montant Total:"
        worksheet['E4'] = f"{montant_total:,.0f} FCFA"
        
        for cell_ref in ['A4', 'D4']:
            worksheet[cell_ref].font = info_font
            worksheet[cell_ref].alignment = info_align
        
        for cell_ref in ['B4', 'E4']:
            worksheet[cell_ref].font = data_font
            worksheet[cell_ref].alignment = info_align
        
        worksheet['A5'] = "Nombre de lignes:"
        worksheet['B5'] = len(df)
        worksheet['A5'].font = info_font
        worksheet['B5'].font = data_font
        worksheet.row_dimensions[5].height = 18
        
        worksheet.row_dimensions[6].height = 5
        
        # EN-TÊTES DE COLONNES
        row_header = 7
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=row_header, column=col_idx)
            cell.value = col_name
            cell.fill = col_header_fill
            cell.font = col_header_font
            cell.alignment = col_header_align
            cell.border = thin_border
        
        worksheet.row_dimensions[row_header].height = 30
        
        # DONNÉES
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=row_header + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border
                
                col_name = df.columns[col_idx - 1]
                if col_name in ['MONTANT', 'MONTANT_']:
                    cell.alignment = align_right
                    if value:
                        cell.number_format = '#,##0'
                elif col_name in ['LIBELLE COMPLEMENTAIRE', 'LIB COMPLEMENTAIRE']:
                    cell.alignment = align_left_wrap
                else:
                    cell.alignment = align_left
        
        # LARGEURS
        col_widths = {
            'CODE AGENCE': 15,
            'COMPTE DE CHARGES': 20,
            'SENS': 10,
            'MONTANT': 18,
            'CODE PAYT Lib 1-5': 18,
            'CODE CHARGE Lib 6-10': 18,
            'TYPE DEP Lib 11': 15,
            'MATR OBJ Lib 12-19': 18,
            'LIBELLE COMPLEMENTAIRE': 50,
            'CODE AG': 12,
            'SENS_': 10,
            'MONTANT_': 18,
            'CODE FOURNISSEUR': 18,
            'FOURNISSEUR': 25,
            'CONTREPARTIE': 20,
            'LIB COMPLEMENTAIRE': 30,
            'IDENTIFIANT': 20
        }
        
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            width = col_widths.get(col_name, 15)
            worksheet.column_dimensions[col_letter].width = width
        
        # RÉCAPITULATIF
        derniere_ligne = row_header + len(df)
        ligne_recap = derniere_ligne + 2
        
        recap_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        recap_font = Font(bold=True, size=11, name='Calibri')
        recap_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        worksheet.cell(row=ligne_recap, column=1, value="Date d'édition:")
        worksheet.cell(row=ligne_recap, column=2, value=current_date)
        worksheet.cell(row=ligne_recap, column=4, value="MONTANT TOTAL:")
        worksheet.cell(row=ligne_recap, column=5, value=f"{montant_total:,.0f} FCFA")
        
        for col in [1, 2, 4, 5]:
            cell = worksheet.cell(row=ligne_recap, column=col)
            cell.font = recap_font
            cell.fill = recap_fill
            cell.border = recap_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.row_dimensions[ligne_recap].height = 25
    
    output.seek(0)
    return output


def page_generation_fichiers():
    """
    Page Streamlit pour la génération des pièces comptables
    """
    st.markdown("## ⚙️ Génération des Pièces Comptables")
    st.markdown("---")
    
    df_central = st.session_state.df_central
    
    if 'DATE' not in df_central.columns or len(df_central) == 0:
        st.warning("⚠️ Aucune donnée dans la base centrale. Importez d'abord des factures.")
        return
    
    # Tabs BT et HT
    tab_bt, tab_ht = st.tabs(["🔌 Basse Tension (BT)", "⚡ Haute Tension (HT)"])
    
    with tab_bt:
        st.markdown("### 📋 Génération Pièce Comptable BT")
        
        # Sélection période
        periodes_disponibles = sorted(df_central['DATE'].dropna().unique().tolist(), reverse=True)
        
        if len(periodes_disponibles) == 0:
            st.warning("⚠️ Aucune période disponible")
        else:
            periode_bt = st.selectbox(
                "📅 Sélectionner la période",
                periodes_disponibles,
                key="periode_gen_bt"
            )
            
            if periode_bt:
                st.info(f"📊 Période sélectionnée : **{periode_bt}**")
                
                # Générer la pièce
                piece_bt = generer_piece_comptable(df_central, periode_bt)
                
                if len(piece_bt) == 0:
                    st.warning(f"⚠️ Aucune donnée pour la période {periode_bt}")
                else:
                    # Statistiques
                    col_stat1, col_stat2, col_stat3 = st.columns(3)
                    
                    with col_stat1:
                        st.metric("📝 Lignes", len(piece_bt))
                    with col_stat2:
                        total = piece_bt['MONTANT'].sum()
                        st.metric("💰 Total", f"{total:,.0f} FCFA")
                    with col_stat3:
                        nb_comp = (piece_bt['LIBELLE COMPLEMENTAIRE'].str.contains('COMPLEMENTAIRE', na=False)).sum()
                        st.metric("🔄 Complémentaires", nb_comp)
                    
                    st.markdown("---")
                    
                    # Aperçu
                    with st.expander("👁️ Aperçu de la pièce comptable BT"):
                        st.dataframe(piece_bt, use_container_width=True, height=400)
                    
                    st.markdown("---")
                    
                    # Actions
                    col1, col2, col3 = st.columns(3)
                    
                    with col2:
                        # Export Excel
                        excel_data = export_piece_comptable_excel(piece_bt, type_tension="BT")
                        
                        st.download_button(
                            "📥 Télécharger Pièce BT",
                            data=excel_data,
                            file_name=f"Piece_Comptable_BT_{periode_bt.replace('/', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
    
    with tab_ht:
        st.markdown("### 📋 Génération Pièce Comptable HT")
        
        # Sélection période
        periodes_disponibles = sorted(df_central['DATE'].dropna().unique().tolist(), reverse=True)
        
        if len(periodes_disponibles) == 0:
            st.warning("⚠️ Aucune période disponible")
        else:
            periode_ht = st.selectbox(
                "📅 Sélectionner la période",
                periodes_disponibles,
                key="periode_gen_ht"
            )
            
            if periode_ht:
                st.info(f"📊 Période sélectionnée : **{periode_ht}**")
                
                # Générer la pièce
                piece_ht = generer_piece_comptable(df_central, periode_ht)
                
                if len(piece_ht) == 0:
                    st.warning(f"⚠️ Aucune donnée pour la période {periode_ht}")
                else:
                    # Statistiques
                    col_stat1, col_stat2, col_stat3 = st.columns(3)
                    
                    with col_stat1:
                        st.metric("📝 Lignes", len(piece_ht))
                    with col_stat2:
                        total = piece_ht['MONTANT'].sum()
                        st.metric("💰 Total", f"{total:,.0f} FCFA")
                    with col_stat3:
                        nb_comp = (piece_ht['LIBELLE COMPLEMENTAIRE'].str.contains('COMPLEMENTAIRE', na=False)).sum()
                        st.metric("🔄 Complémentaires", nb_comp)
                    
                    st.markdown("---")
                    
                    # Aperçu
                    with st.expander("👁️ Aperçu de la pièce comptable HT"):
                        st.dataframe(piece_ht, use_container_width=True, height=400)
                    
                    st.markdown("---")
                    
                    # Actions
                    col1, col2, col3 = st.columns(3)
                    
                    with col2:
                        # Export Excel
                        excel_data = export_piece_comptable_excel(piece_ht, type_tension="HT")
                        
                        st.download_button(
                            "📥 Télécharger Pièce HT",
                            data=excel_data,
                            file_name=f"Piece_Comptable_HT_{periode_ht.replace('/', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
