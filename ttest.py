import streamlit as st
import pandas as pd
import io
import pickle
import os
from datetime import datetime
import plotly.graph_objects as go

st.set_page_config(
    page_title="Gestion Factures - Historique par Ligne",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personnalisé
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .main-title {
        color: white;
        font-size: 2.5rem;
        font-weight: bold;
        margin: 0;
        text-align: center;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        text-align: center;
    }
</style>
""", unsafe_allow_html=True)

# Fichiers
FICHIER_CENTRAL = "Base_Centrale.xlsx"
FICHIER_TEMPLATE_BT = "FACTURAT_ELECTRICITE_BT.xlsx"
FICHIER_TEMPLATE_HT = "FACTURAT_ELECTRICITE_HT.xlsx"
SAVE_FILE_CENTRAL = "data_centrale.pkl"


def normaliser_identifiant(valeur):
    """Normalise un identifiant : supprime .0, espaces, met en majuscules"""
    if pd.isna(valeur):
        return ''
    valeur_str = str(valeur).strip()

    try:
        valeur_float = float(valeur_str)
        if valeur_float.is_integer():
            return str(int(valeur_float)).upper()
        else : 
            return valeur_str.upper()
    except:
        return valeur_str.upper()


# Fonctions de chargement
def load_central():
    if os.path.exists(SAVE_FILE_CENTRAL):
        with open(SAVE_FILE_CENTRAL, 'rb') as f:
            df = pickle.load(f)
    elif os.path.exists(FICHIER_CENTRAL):
        df = pd.read_excel(FICHIER_CENTRAL)
    
        for col in ['MONTANT', 'CONSO', 'DATE']:
            if col not in df.columns:
                df[col] = None
    else:
        st.error(f"❌ Fichier central '{FICHIER_CENTRAL}' introuvable !")
        st.stop()

    if 'IDENTIFIANT' in df.columns:
        df["IDENTIFIANT"] = df["IDENTIFIANT"].apply(normaliser_identifiant)

    return df


def load_template_bt():
    """Charge le template BT avec sa structure"""
    if os.path.exists(FICHIER_TEMPLATE_BT):
        df = pd.read_excel(FICHIER_TEMPLATE_BT)
        if 'IDENTIFIANT' in df.columns:
            df["IDENTIFIANT"] = df["IDENTIFIANT"].apply(normaliser_identifiant)
        return df
    return None


def load_template_ht():
    """Charge le template HT avec sa structure"""
    if os.path.exists(FICHIER_TEMPLATE_HT):
        df = pd.read_excel(FICHIER_TEMPLATE_HT)
        if 'IDENTIFIANT' in df.columns:
            df["IDENTIFIANT"] = df["IDENTIFIANT"].apply(normaliser_identifiant)
        return df
    return None


def save_central(df):
    with open(SAVE_FILE_CENTRAL, 'wb') as f:
        pickle.dump(df, f)


def export_base_centrale(df):
    """Génère un fichier Excel avec le style pour la Base Centrale (vert)"""
    output = io.BytesIO()
    
    # Remplacer les valeurs NA par des chaînes vides
    df = df.fillna('')
    
    # Couleurs vertes pour la base centrale
    header_color = '375623'  # Vert foncé
    col_header_color = '70AD47'  # Vert
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet('Base_Centrale')
        
        # Supprimer la feuille par défaut
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        # === STYLES ===
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        header_font = Font(bold=True, size=16, color='FFFFFF', name='Calibri')
        header_align = Alignment(horizontal='center', vertical='center')
        
        subtitle_font = Font(bold=True, size=12, color=header_color, name='Calibri')
        subtitle_align = Alignment(horizontal='center', vertical='center')
        
        info_font = Font(bold=True, size=10, name='Calibri')
        info_align = Alignment(horizontal='left', vertical='center')
        
        col_header_fill = PatternFill(start_color=col_header_color, end_color=col_header_color, fill_type='solid')
        col_header_font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
        col_header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        data_align = Alignment(horizontal='left', vertical='center')
        data_font = Font(size=10, name='Calibri')
        
        # === CONSTRUCTION ===
        nb_cols = len(df.columns)
        
        # Ligne 1-2: EN-TÊTE PRINCIPAL
        worksheet.merge_cells(f'A1:{get_column_letter(nb_cols)}2')
        cell = worksheet['A1']
        cell.value = "📊 BASE CENTRALE"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        worksheet.row_dimensions[1].height = 25
        worksheet.row_dimensions[2].height = 25
        
        # Ligne 3: SOUS-TITRE
        worksheet.merge_cells(f'A3:{get_column_letter(nb_cols)}3')
        cell = worksheet['A3']
        cell.value = "HISTORIQUE COMPLET DES FACTURES ÉLECTRIQUES"
        cell.font = subtitle_font
        cell.alignment = subtitle_align
        worksheet.row_dimensions[3].height = 20
        
        # Ligne 4: Date et statistiques
        current_date = datetime.now().strftime('%d/%m/%Y %H:%M')
        montant_total = df['MONTANT'].sum() if 'MONTANT' in df.columns else 0
        
        worksheet['A4'] = "Date d'export:"
        worksheet['B4'] = current_date
        worksheet['D4'] = "Montant Total"
        worksheet['E4'] = f"{montant_total:,.0f} FCFA"
        
        for cell_ref in ['A4', 'D4']:
            worksheet[cell_ref].font = info_font
            worksheet[cell_ref].alignment = info_align
        
        for cell_ref in ['B4', 'E4']:
            worksheet[cell_ref].font = data_font
            worksheet[cell_ref].alignment = info_align
        
        # Ligne 5: Nombre de lignes et sites
        sites_uniques = df['IDENTIFIANT'].nunique() if 'IDENTIFIANT' in df.columns else 0
        
        worksheet['A5'] = "Nombre de lignes:"
        worksheet['B5'] = len(df)
        worksheet['D5'] = "Sites uniques:"
        worksheet['E5'] = sites_uniques
        
        for cell_ref in ['A5', 'D5']:
            worksheet[cell_ref].font = info_font
            worksheet[cell_ref].alignment = info_align
        
        for cell_ref in ['B5', 'E5']:
            worksheet[cell_ref].font = data_font
            worksheet[cell_ref].alignment = info_align
        
        worksheet.row_dimensions[5].height = 18
        
        # Ligne 6: Vide
        worksheet.row_dimensions[6].height = 5
        
        # Ligne 7: EN-TÊTES DE COLONNES
        row_header = 7
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=row_header, column=col_idx)
            cell.value = col_name
            cell.fill = col_header_fill
            cell.font = col_header_font
            cell.alignment = col_header_align
            cell.border = thin_border
        
        worksheet.row_dimensions[row_header].height = 30
        
        # Lignes de données
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=row_header + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                
                # Formater les montants
                if df.columns[col_idx - 1] == 'MONTANT' and value:
                    cell.number_format = '#,##0'
        
        # === AJUSTER LARGEURS ===
        col_widths = {
            'UC': 10,
            'CODE AGCE': 12,
            'SITES': 25,
            'CORRESPONDANCE': 20,
            'IDENTIFIANT': 15,
            'REFERENCE': 15,
            'TENSION': 10,
            'MONTANT': 12,
            'CONSO': 12,
            'DATE': 12
        }
        
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            width = col_widths.get(col_name, 15)
            worksheet.column_dimensions[col_letter].width = width
    
    output.seek(0)
    return output


def export_factures_cie(df, type_tension="BT"):
    """Génère un fichier Excel avec le style CIE professionnel pour BT (bleu) ou HT (orange)"""
    output = io.BytesIO()

    # Remplacer les valeurs NA par des chaînes vides
    df = df.fillna('')
    
    # Définir les couleurs selon le type
    if type_tension == "BT":
        header_color = '2F5597'  # Bleu
        col_header_color = '4472C4'  # Bleu clair
        subtitle_color = '2F5597'  # Bleu
    else:  # HT
        header_color = 'C65911'  # Orange foncé
        col_header_color = 'ED7D31'  # Orange
        subtitle_color = 'C65911'  # Orange foncé
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet('Factures')
        
        # Supprimer la feuille par défaut
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        # === STYLES ===
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        header_font = Font(bold=True, size=16, color='FFFFFF', name='Calibri')
        header_align = Alignment(horizontal='center', vertical='center')
        
        subtitle_font = Font(bold=True, size=12, color=subtitle_color, name='Calibri')
        subtitle_align = Alignment(horizontal='center', vertical='center')
        
        info_font = Font(bold=True, size=10, name='Calibri')
        info_align = Alignment(horizontal='left', vertical='center')
        
        col_header_fill = PatternFill(start_color=col_header_color, end_color=col_header_color, fill_type='solid')
        col_header_font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
        col_header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        data_align = Alignment(horizontal='left', vertical='center')
        data_font = Font(size=10, name='Calibri')
        
        # === CONSTRUCTION ===
        nb_cols = len(df.columns)
        
        # Ligne 1-2: EN-TÊTE PRINCIPAL
        worksheet.merge_cells(f'A1:{get_column_letter(nb_cols)}2')
        cell = worksheet['A1']
        cell.value = "⚡ COMPAGNIE IVOIRIENNE D'ÉLECTRICITÉ (CIE)"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        worksheet.row_dimensions[1].height = 25
        worksheet.row_dimensions[2].height = 25
        
        # Ligne 3: SOUS-TITRE
        worksheet.merge_cells(f'A3:{get_column_letter(nb_cols)}3')
        cell = worksheet['A3']
        tension_txt = "BASSE TENSION (BT)" if type_tension == "BT" else "HAUTE TENSION (HT)"
        cell.value = f"FACTURES {tension_txt} - RAPPORT MENSUEL"
        cell.font = subtitle_font
        cell.alignment = subtitle_align
        worksheet.row_dimensions[3].height = 20
        
        # Ligne 4: Date et Montant
        current_date = datetime.now().strftime('%d/%m/%Y %H:%M')
        montant_total = df['MONTANT'].sum() if 'MONTANT' in df.columns else 0
        
        worksheet['A4'] = "Date d'édition:"
        worksheet['B4'] = current_date
        worksheet['D4'] = "Montant"
        worksheet['E4'] = f"{montant_total:,.0f} FCFA"
        
        for cell_ref in ['A4', 'D4']:
            worksheet[cell_ref].font = info_font
            worksheet[cell_ref].alignment = info_align
        
        for cell_ref in ['B4', 'E4']:
            worksheet[cell_ref].font = data_font
            worksheet[cell_ref].alignment = info_align
        
        # Ligne 5: Nombre de factures
        worksheet['A5'] = "Nombre de factures:"
        worksheet['B5'] = len(df)
        worksheet['A5'].font = info_font
        worksheet['B5'].font = data_font
        worksheet.row_dimensions[5].height = 18
        
        # Ligne 6: Vide
        worksheet.row_dimensions[6].height = 5
        
        # Ligne 7: EN-TÊTES DE COLONNES
        row_header = 7
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=row_header, column=col_idx)
            cell.value = col_name
            cell.fill = col_header_fill
            cell.font = col_header_font
            cell.alignment = col_header_align
            cell.border = thin_border
        
        worksheet.row_dimensions[row_header].height = 30
        
        # Lignes de données
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=row_header + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                
                # Formater les montants
                if df.columns[col_idx - 1] == 'MONTANT' and value:
                    cell.number_format = '#,##0'
        
        # === SECTION SIGNATURES ===
        derniere_ligne_data = row_header + len(df)
        ligne_vide = derniere_ligne_data + 2
        ligne_titre_sig = ligne_vide + 1
        
        # Titre "SIGNATURES ET APPROBATIONS"
        worksheet.merge_cells(f'A{ligne_titre_sig}:{get_column_letter(nb_cols)}{ligne_titre_sig}')
        cell = worksheet[f'A{ligne_titre_sig}']
        cell.value = "SIGNATURES ET APPROBATIONS"
        cell.font = Font(bold=True, size=12, color=subtitle_color, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        worksheet.row_dimensions[ligne_titre_sig].height = 25
        
        # Ligne des labels
        ligne_labels = ligne_titre_sig + 2
        
        # Calculer l'espacement
        col_prep = 1
        col_verif = max(4, nb_cols // 4)
        col_valid = max(7, nb_cols // 2)
        col_appro = max(10, 3 * nb_cols // 4)
        
        # Labels
        labels = [
            (col_prep, "Préparé par:"),
            (col_verif, "Vérifié par:"),
            (col_valid, "Validé par:"),
            (col_appro, "Approuvé par:")
        ]
        
        for col, label in labels:
            cell = worksheet.cell(row=ligne_labels, column=col)
            cell.value = label
            cell.font = Font(bold=True, size=10, name='Calibri')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(bottom=Side(style='medium', color='000000'))
        
        # Rectangles pour signatures
        ligne_sig_start = ligne_labels + 1
        ligne_sig_end = ligne_sig_start + 3
        
        rectangles = [
            (col_prep, col_prep + 2),
            (col_verif, col_verif + 2),
            (col_valid, col_valid + 2),
            (col_appro, col_appro + 2)
        ]
        
        for start_col, end_col in rectangles:
            worksheet.merge_cells(
                start_row=ligne_sig_start,
                start_column=start_col,
                end_row=ligne_sig_end,
                end_column=end_col
            )
            
            cell = worksheet.cell(row=ligne_sig_start, column=start_col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in range(ligne_sig_start, ligne_sig_end + 1):
            worksheet.row_dimensions[row].height = 20
        
        # Lignes Nom: et Date:
        ligne_nom = ligne_sig_end + 1
        ligne_date = ligne_nom + 1
        
        for col, _ in labels:
            worksheet.cell(row=ligne_nom, column=col, value="Nom:")
            worksheet.cell(row=ligne_nom, column=col).font = Font(size=9, name='Calibri')
            
            worksheet.cell(row=ligne_date, column=col, value="Date:")
            worksheet.cell(row=ligne_date, column=col).font = Font(size=9, name='Calibri')
        
        # === AJUSTER LARGEURS ===
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 15
    
    output.seek(0)
    return output


# Initialisation
if 'df_central' not in st.session_state:
    st.session_state.df_central = load_central()

# Header
st.markdown("""
<div class="main-header">
    <h1 class="main-title" style="color: #f0f0f0">📊 Gestion Centralisée des Factures</h1>
    <p class="main-subtitle" style="color: #f0f0f0; text-align: center; margin-top: 0.5rem;">
        Système avec historique mensuel
    </p>
</div>
""", unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.markdown("### 📋 Navigation")
    
    page = st.radio(
        "Menu principal",
        ["📊 Base Centrale", "🔄 Import Factures BT", "🔄 Import Factures HT", "📈 Statistiques", "⚙️ Génération Fichiers"]
    )
    
    st.markdown("---")
    st.markdown("### 📊 Informations")
    
    df_central = st.session_state.df_central
    st.metric("📝 Lignes totales", len(df_central[df_central['DATE'].notna() & (df_central['DATE'] != '')]))
    
    if 'DATE' in df_central.columns:
        periodes = df_central['DATE'].dropna().nunique()
        st.metric("📅 Périodes", periodes)

# CONTENU PRINCIPAL
if page == "📊 Base Centrale":
    st.markdown("## 📊 Base Centrale - Historique Complet")
    st.markdown("---")
    
    df_central = st.session_state.df_central
    
    # Statistiques
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: #667eea;">📝</h3>
            <h2>{len(df_central[df_central['DATE'].notna() & (df_central['DATE'] != '')])}</h2> 
            <p style="color: #666;">Lignes totales</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        sites_uniques = df_central['IDENTIFIANT'].nunique() if 'IDENTIFIANT' in df_central.columns else 0
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: #667eea;">🏢</h3>
            <h2>{sites_uniques}</h2>
            <p style="color: #666;">Sites uniques</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        periodes = df_central['DATE'].dropna().nunique() if 'DATE' in df_central.columns else 0
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: #667eea;">📅</h3>
            <h2>{periodes}</h2>
            <p style="color: #666;">Périodes</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        total = df_central['MONTANT'].sum() if 'MONTANT' in df_central.columns else 0
        st.markdown(f"""
        <div class="metric-card">
            <h3 style="color: #667eea;">💰</h3>
            <h2>{total/1000:.0f}K</h2>
            <p style="color: #666;">Total FCFA</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Filtres
    col_f1, col_f2, col_f3 = st.columns(3)
    
    with col_f1:
        if 'UC' in df_central.columns:
            ucs = ['Tous'] + sorted(df_central['UC'].dropna().unique().tolist())
            uc_filter = st.selectbox("Filtrer par UC", ucs)
        else:
            uc_filter = 'Tous'
    
    with col_f2:
        if 'DATE' in df_central.columns:
            dates = ['Tous'] + sorted(df_central['DATE'].dropna().unique().tolist(), reverse=True)
            date_filter = st.selectbox("Filtrer par DATE", dates)
        else:
            date_filter = 'Tous'
    
    with col_f3:
        if 'TENSION' in df_central.columns:
            tensions = ['Tous'] + sorted(df_central['TENSION'].dropna().unique().tolist())
            tension_filter = st.selectbox("Filtrer par TENSION", tensions)
        else:
            tension_filter = 'Tous'
    
    # Appliquer les filtres
    df_filtered = df_central.copy()

    if 'DATE' in df_filtered.columns : 
        df_filtered = df_filtered[df_filtered['DATE'].notna() & (df_filtered['DATE'] != '')]

    if uc_filter != 'Tous' and 'UC' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['UC'] == uc_filter]
    if date_filter != 'Tous' and 'DATE' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['DATE'] == date_filter]
    if tension_filter != 'Tous' and 'TENSION' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['TENSION'] == tension_filter]
    
    st.markdown(f"### 📋 Données filtrées ({len(df_filtered)} ligne(s))")
    
    # Tableau
    edited_df = st.data_editor(
        df_filtered,
        use_container_width=True,
        num_rows="dynamic",
        height=500,
        key="editor_central"
    )
    
    # Actions
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("💾 Sauvegarder", type="primary", use_container_width=True):
            for idx in edited_df.index:
                if idx in st.session_state.df_central.index:
                    st.session_state.df_central.loc[idx] = edited_df.loc[idx]
            save_central(st.session_state.df_central)
            st.success("✅ Base centrale sauvegardée !")
            st.rerun()
    
    with col2:
        # Export base centrale avec design vert
        excel_data = export_base_centrale(df_filtered)
        
        st.download_button(
            "📥 Exporter Excel",
            data=excel_data,
            file_name=f"Base_Centrale_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_central"
        )
    
    with col3:
        if st.button("🔄 Actualiser", use_container_width=True):
            st.rerun()

elif page == "🔄 Import Factures BT":
    st.markdown("## 🔄 Import Factures - Basse Tension (BT)")
    st.markdown("---")
    
    st.markdown("""
    <div style='background: linear-gradient(135deg, #667eea, #764ba2); 
                color: white; 
                padding: 1.5rem; 
                border-radius: 10px;
                margin: 1rem 0;'>
        <h3 style='margin: 0;'>🔌 BASSE TENSION</h3>
        <p style='margin: 0.5rem 0 0 0;'>Import mensuel - Ajout de nouvelles lignes dans l'historique</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.info("""
    📌 **Configuration BT** :
    - Clé facture : **Référence Contrat**
    - Clé base centrale : **IDENTIFIANT**
    - Données : **Montant facture TTC**, **KWH Facturé**, **Période Facture sur date fact**
    
    💡 Pour chaque ligne trouvée, une **nouvelle ligne** sera ajoutée dans la base centrale avec les données du mois.
    """)
    
    # Upload fichier
    fichier_bt = st.file_uploader(
        "Sélectionnez le fichier de factures BT",
        type=['xlsx', 'xls'],
        key="upload_bt"
    )
    
    if fichier_bt:
        try:
            # Configuration des colonnes
            cle_facture = "Référence Contrat"
            montant_col = "Montant facture TTC"
            conso_col = "KWH Facturé"
            caract_col = "Période Facture sur date fact"

            df_bt = pd.read_excel(fichier_bt)
            
            # NORMALISATION COMPLÈTE DE LA PÉRIODE (CORRECTION CRITIQUE)
            df_bt[caract_col] = df_bt[caract_col].astype(str).str.strip()
            # Enlever les .0 si présents
            df_bt[caract_col] = df_bt[caract_col].str.replace('.0', '', regex=False)
            # S'assurer que c'est au format YYYYMM (6 chiffres)
            df_bt[caract_col] = df_bt[caract_col].str.zfill(6)
            # Convertir en MM/YYYY avec zéro initial
            df_bt[caract_col] = df_bt[caract_col].str[4:6] + '/' + df_bt[caract_col].str[:4]
            
            # Normaliser les identifiants
            if cle_facture in df_bt.columns:
                df_bt[cle_facture] = df_bt[cle_facture].apply(normaliser_identifiant)
                
            st.success(f"✅ Fichier chargé : {len(df_bt)} ligne(s)")
            
            # Vérifications
            colonnes_manquantes = []
            for col in [cle_facture, montant_col, caract_col]:
                if col not in df_bt.columns:
                    colonnes_manquantes.append(col)
            
            if colonnes_manquantes:
                st.error(f"❌ Colonnes manquantes : {', '.join(colonnes_manquantes)}")
                st.info(f"📋 Colonnes disponibles : {', '.join(df_bt.columns)}")
            else:
                # Récupérer la période
                periode_bt = df_bt[caract_col].dropna().unique()
                if len(periode_bt) > 0:
                    periode_bt = str(periode_bt[0])
                    st.success(f"✅ Période BT détectée : **{periode_bt}**")
                else:
                    periode_bt = ""
                    st.warning("⚠️ Aucune période détectée")
                
                # Aperçu
                with st.expander("👁️ Aperçu du fichier BT"):
                    cols_to_show = [cle_facture, montant_col, caract_col]
                    if conso_col in df_bt.columns:
                        cols_to_show.insert(2, conso_col)
                    st.dataframe(df_bt[cols_to_show].head(10), use_container_width=True)
                
                st.markdown("---")
                
                # Bouton import
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    if st.button("🔄 LANCER L'IMPORT BT", type="primary", use_container_width=True):
                        with st.spinner("⏳ Import BT en cours..."):
                            df_central = st.session_state.df_central.copy()
                            df_template_bt = load_template_bt()
                            
                            if df_template_bt is None:
                                st.error("❌ Fichier template BT introuvable !")
                            else:
                                nouvelles_lignes = []
                                nb_ajouts = 0
                                
                                for _, row_facture in df_bt.iterrows():
                                    ref_contrat = normaliser_identifiant(row_facture[cle_facture])
                                    
                                    # Chercher dans la base centrale pour récupérer les infos du site
                                    ligne_centrale = df_central[df_central['IDENTIFIANT'] == ref_contrat]
                                    
                                    if not ligne_centrale.empty:
                                        # Créer une nouvelle ligne avec les colonnes de la base centrale
                                        nouvelle_ligne = pd.Series({
                                            'UC': ligne_centrale.iloc[0]['UC'] if 'UC' in ligne_centrale.columns else '',
                                            'CODE AGCE': ligne_centrale.iloc[0]['CODE AGCE'] if 'CODE AGCE' in ligne_centrale.columns else '',
                                            'SITES': ligne_centrale.iloc[0]['SITES'] if 'SITES' in ligne_centrale.columns else '',
                                            'CORRESPONDANCE': ligne_centrale.iloc[0]['CORRESPONDANCE'] if 'CORRESPONDANCE' in ligne_centrale.columns else '',
                                            'IDENTIFIANT': ref_contrat,
                                            'REFERENCE': ligne_centrale.iloc[0]['REFERENCE'] if 'REFERENCE' in ligne_centrale.columns else '',
                                            'TENSION': ligne_centrale.iloc[0]['TENSION'] if 'TENSION' in ligne_centrale.columns else '',
                                            'MONTANT': row_facture[montant_col],
                                            'CONSO': row_facture.get(conso_col, None) if conso_col in df_bt.columns else None,
                                            'DATE': periode_bt
                                        })
                                        
                                        nouvelles_lignes.append(nouvelle_ligne)
                                        nb_ajouts += 1
                            
                            if nouvelles_lignes:
                                # Ajouter les nouvelles lignes à la base centrale
                                df_nouvelles = pd.DataFrame(nouvelles_lignes)
                                df_central = pd.concat([df_central, df_nouvelles], ignore_index=True)
                                
                                # Compter avant suppression des doublons
                                nb_avant = len(df_central)
                                
                                # Supprimer les doublons (IDENTIFIANT + DATE)
                                df_central = df_central.drop_duplicates(subset=['IDENTIFIANT', 'DATE'], keep='first')
                                
                                # Calculer le nombre de doublons supprimés
                                nb_apres = len(df_central)
                                nb_doublons = nb_avant - nb_apres
                                nb_ajoutes = len(nouvelles_lignes) - nb_doublons
                                
                                # Sauvegarder
                                st.session_state.df_central = df_central
                                save_central(df_central)
                                
                                # Résultats
                                st.markdown("---")
                                st.success(f"🎉 Import BT terminé : {nb_ajoutes} ligne(s) ajoutée(s) !")
                                
                                col_r1, col_r2, col_r3 = st.columns(3)
                                with col_r1:
                                    st.metric("✅ Lignes ajoutées", nb_ajoutes)
                                with col_r2:
                                    st.metric("📊 Total lignes", len(df_central))
                                with col_r3:
                                    st.metric("📅 Période", periode_bt)
                                
                                if nb_doublons > 0:
                                    st.info(f"ℹ️ {nb_doublons} doublon(s) supprimé(s) (IDENTIFIANT + DATE déjà existants)")
                                
                                st.balloons()
                            else:
                                st.warning("⚠️ Aucune correspondance trouvée")
        
        except Exception as e:
            st.error(f"❌ Erreur : {str(e)}")
            st.exception(e)

elif page == "🔄 Import Factures HT":
    st.markdown("## 🔄 Import Factures - Haute Tension (HT)")
    st.markdown("---")
    
    st.markdown("""
    <div style='background: linear-gradient(135deg, #f093fb, #f5576c); 
                color: white; 
                padding: 1.5rem; 
                border-radius: 10px;
                margin: 1rem 0;'>
        <h3 style='margin: 0;'>⚡ HAUTE TENSION</h3>
        <p style='margin: 0.5rem 0 0 0;'>Import mensuel - Ajout de nouvelles lignes dans l'historique</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.info("""
    📌 **Configuration HT** :
    - Clé facture : **refraccord**
    - Clé base centrale : **IDENTIFIANT**
    - Données : **montfact**, **conso**, **Periode_Emission_Bordereau**
    
    💡 Pour chaque ligne trouvée, une **nouvelle ligne** sera ajoutée dans la base centrale avec les données du mois.
    """)
    
    # Upload fichier
    fichier_ht = st.file_uploader(
        "Sélectionnez le fichier de factures HT",
        type=['xlsx', 'xls'],
        key="upload_ht"
    )
    
    if fichier_ht:
        try:
            cle_facture = "refraccord"
            montant_col = "montfact"
            conso_col = "conso"
            caract_col = "Periode_Emission_Bordereau"

            df_ht = pd.read_excel(fichier_ht)
            
            # Normalisation du refraccord
            df_ht[cle_facture] = df_ht[cle_facture].apply(lambda x : str(x).replace(".0",""))
            df_ht[cle_facture] = df_ht[cle_facture].astype('string')
     
            # Normalisation de la période
            df_ht[caract_col] = df_ht[caract_col].astype(str).str.replace(".0", "").str.zfill(6)
            df_ht[caract_col] = df_ht[caract_col].str[:2] + "/" + df_ht[caract_col].str[2:]

            if cle_facture in df_ht.columns:
                df_ht[cle_facture] = df_ht[cle_facture].apply(normaliser_identifiant)

            st.success(f"✅ Fichier chargé : {len(df_ht)} ligne(s)")
            
            # Vérifications
            colonnes_manquantes = []
            for col in [cle_facture, montant_col, caract_col]:
                if col not in df_ht.columns:
                    colonnes_manquantes.append(col)
            
            if colonnes_manquantes:
                st.error(f"❌ Colonnes manquantes : {', '.join(colonnes_manquantes)}")
                st.info(f"📋 Colonnes disponibles : {', '.join(df_ht.columns)}")
            else:
                # Récupérer la période
                periode_ht = df_ht[caract_col].dropna().unique()
                if len(periode_ht) > 0:
                    periode_ht = str(periode_ht[0])
                    st.success(f"✅ Période HT détectée : **{periode_ht}**")
                else:
                    periode_ht = ""
                    st.warning("⚠️ Aucune période détectée")
                
                # Aperçu
                with st.expander("👁️ Aperçu du fichier HT"):
                    cols_to_show = [cle_facture, montant_col, caract_col]
                    if conso_col in df_ht.columns:
                        cols_to_show.insert(2, conso_col)
                    st.dataframe(df_ht[cols_to_show].head(10), use_container_width=True)
                
                st.markdown("---")
                
                # Bouton import
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    if st.button("🔄 LANCER L'IMPORT HT", type="primary", use_container_width=True):
                        with st.spinner("⏳ Import HT en cours..."):
                            df_central = st.session_state.df_central.copy()
                            df_template_ht = load_template_ht()
                            
                            if df_template_ht is None:
                                st.error("❌ Fichier template HT introuvable !")
                            else:
                                nouvelles_lignes = []
                                nb_ajouts = 0
                                
                                for _, row_facture in df_ht.iterrows():
                                    refraccord = normaliser_identifiant(row_facture[cle_facture])
                                    
                                    # Chercher dans la base centrale pour récupérer les infos du site
                                    ligne_centrale = df_central[df_central['IDENTIFIANT'] == refraccord]
                                    
                                    if not ligne_centrale.empty:
                                        # Créer une nouvelle ligne avec les colonnes de la base centrale
                                        nouvelle_ligne = pd.Series({
                                            'UC': ligne_centrale.iloc[0]['UC'] if 'UC' in ligne_centrale.columns else '',
                                            'CODE AGCE': ligne_centrale.iloc[0]['CODE AGCE'] if 'CODE AGCE' in ligne_centrale.columns else '',
                                            'SITES': ligne_centrale.iloc[0]['SITES'] if 'SITES' in ligne_centrale.columns else '',
                                            'CORRESPONDANCE': ligne_centrale.iloc[0]['CORRESPONDANCE'] if 'CORRESPONDANCE' in ligne_centrale.columns else '',
                                            'IDENTIFIANT': refraccord,
                                            'REFERENCE': ligne_centrale.iloc[0]['REFERENCE'] if 'REFERENCE' in ligne_centrale.columns else '',
                                            'TENSION': ligne_centrale.iloc[0]['TENSION'] if 'TENSION' in ligne_centrale.columns else '',
                                            'MONTANT': row_facture[montant_col],
                                            'CONSO': row_facture.get(conso_col, None) if conso_col in df_ht.columns else None,
                                            'DATE': periode_ht
                                        })
                                        
                                        nouvelles_lignes.append(nouvelle_ligne)
                                        nb_ajouts += 1
                            
                            if nouvelles_lignes:
                                # Ajouter les nouvelles lignes à la base centrale
                                df_nouvelles = pd.DataFrame(nouvelles_lignes)
                                df_central = pd.concat([df_central, df_nouvelles], ignore_index=True)
                                
                                # Compter avant suppression des doublons
                                nb_avant = len(df_central)
                                
                                # Supprimer les doublons (IDENTIFIANT + DATE)
                                df_central = df_central.drop_duplicates(subset=['IDENTIFIANT', 'DATE'], keep='first')
                                
                                # Calculer le nombre de doublons supprimés
                                nb_apres = len(df_central)
                                nb_doublons = nb_avant - nb_apres
                                nb_ajoutes = len(nouvelles_lignes) - nb_doublons
                                
                                # Sauvegarder
                                st.session_state.df_central = df_central
                                save_central(df_central)
                                
                                # Résultats
                                st.markdown("---")
                                st.success(f"🎉 Import HT terminé : {nb_ajoutes} ligne(s) ajoutée(s) !")
                                
                                col_r1, col_r2, col_r3 = st.columns(3)
                                with col_r1:
                                    st.metric("✅ Lignes ajoutées", nb_ajoutes)
                                with col_r2:
                                    st.metric("📊 Total lignes", len(df_central))
                                with col_r3:
                                    st.metric("📅 Période", periode_ht)
                                
                                if nb_doublons > 0:
                                    st.info(f"ℹ️ {nb_doublons} doublon(s) supprimé(s) (IDENTIFIANT + DATE déjà existants)")
                                
                                st.balloons()
                            else:
                                st.warning("⚠️ Aucune correspondance trouvée")
        
        except Exception as e:
            st.error(f"❌ Erreur : {str(e)}")
            st.exception(e)

elif page == "📈 Statistiques":
    st.markdown("## 📈 Statistiques et Évolution")
    st.markdown("---")
    
    df_central = st.session_state.df_central
    
    if 'DATE' not in df_central.columns or df_central['DATE'].isna().all():
        st.warning("⚠️ Aucune période enregistrée. Importez d'abord des factures.")
    else:
        periodes_brutes = sorted(df_central['DATE'].dropna().unique().tolist())
        
        if len(periodes_brutes) == 0:
            st.warning("⚠️ Aucune donnée disponible.")
        else:
            st.success(f"✅ {len(periodes_brutes)} période(s) disponible(s)")
            
            # Filtres
            col_f1, col_f2 = st.columns(2)
            
            with col_f1:
                if 'SITES' in df_central.columns:
                    sites = ['Tous'] + sorted(df_central['SITES'].dropna().unique().tolist())
                    site_filter = st.selectbox("🏢 Filtrer par SITE", sites)
                else:
                    site_filter = 'Tous'
            
            with col_f2:
                type_graphique = st.selectbox(
                    "⚡ Type d'analyse",
                    ["📊 Global (BT + HT)", "🔌 Basse Tension uniquement", "⚡ Haute Tension uniquement"]
                )
            
            # Appliquer les filtres
            df_filtered = df_central.copy()
            
            if site_filter != 'Tous':
                df_filtered = df_filtered[df_filtered['SITES'] == site_filter]
            
            if "Basse Tension" in type_graphique:
                df_filtered = df_filtered[df_filtered['TENSION'] == 'BASSE']
            elif "Haute Tension" in type_graphique:
                df_filtered = df_filtered[df_filtered['TENSION'] == 'HAUTE']
            
            # Grouper par DATE
            df_grouped = df_filtered.groupby('DATE').agg({
                'MONTANT': 'sum',
                'CONSO': 'sum'
            }).reset_index()
            
            df_grouped['DATE_DISPLAY'] = df_grouped['DATE']
            df_grouped = df_grouped.sort_values('DATE')
            
            st.markdown("---")
            
            # GRAPHIQUE MONTANTS
            st.markdown("### 💰 Évolution des Montants")
            
            fig_montant = go.Figure()
            
            fig_montant.add_trace(go.Scatter(
                x=df_grouped['DATE_DISPLAY'],
                y=df_grouped['MONTANT'],
                mode='lines+markers',
                name='Montant',
                line=dict(color='#667eea', width=3),
                marker=dict(size=10, color='#667eea'),
                hovertemplate='<b>%{x}</b><br>Montant: %{y:,.0f} FCFA<extra></extra>'
            ))
            
            fig_montant.update_layout(
                title=f"Évolution des Montants - {type_graphique.split(' ', 1)[1] if site_filter == 'Tous' else site_filter}",
                xaxis_title="Période",
                yaxis_title="Montant (FCFA)",
                hovermode='x unified',
                template='plotly_white',
                height=400
            )
            
            st.plotly_chart(fig_montant, use_container_width=True)
            
            # GRAPHIQUE CONSOMMATIONS
            st.markdown("### ⚡ Évolution des Consommations")
            
            fig_conso = go.Figure()
            
            fig_conso.add_trace(go.Scatter(
                x=df_grouped['DATE_DISPLAY'],
                y=df_grouped['CONSO'],
                mode='lines+markers',
                name='Consommation',
                line=dict(color='#f5576c', width=3),
                marker=dict(size=10, color='#f5576c'),
                hovertemplate='<b>%{x}</b><br>Conso: %{y:,.0f} kWh<extra></extra>'
            ))
            
            fig_conso.update_layout(
                title=f"Évolution des Consommations - {type_graphique.split(' ', 1)[1] if site_filter == 'Tous' else site_filter}",
                xaxis_title="Période",
                yaxis_title="Consommation (kWh)",
                hovermode='x unified',
                template='plotly_white',
                height=400
            )
            
            st.plotly_chart(fig_conso, use_container_width=True)
            
            # GRAPHIQUE COMBINÉ
            st.markdown("### 📊 Vue Combinée (Montant + Consommation)")
            
            fig_combine = go.Figure()
            
            fig_combine.add_trace(go.Bar(
                x=df_grouped['DATE_DISPLAY'],
                y=df_grouped['MONTANT'],
                name='Montant',
                marker_color='#667eea',
                yaxis='y',
                hovertemplate='<b>%{x}</b><br>Montant: %{y:,.0f} FCFA<extra></extra>'
            ))
            
            fig_combine.add_trace(go.Scatter(
                x=df_grouped['DATE_DISPLAY'],
                y=df_grouped['CONSO'],
                name='Consommation',
                line=dict(color='#f5576c', width=3),
                marker=dict(size=10, color='#f5576c'),
                yaxis='y2',
                hovertemplate='<b>%{x}</b><br>Conso: %{y:,.0f} kWh<extra></extra>'
            ))
            
            fig_combine.update_layout(
                title=f"Montant vs Consommation - {type_graphique.split(' ', 1)[1] if site_filter == 'Tous' else site_filter}",
                xaxis_title="Période",
                yaxis=dict(
                    title="Montant (FCFA)",
                    side='left',
                    showgrid=False
                ),
                yaxis2=dict(
                    title="Consommation (kWh)",
                    side='right',
                    overlaying='y',
                    showgrid=False
                ),
                hovermode='x unified',
                template='plotly_white',
                height=500,
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1
                )
            )
            
            st.plotly_chart(fig_combine, use_container_width=True)
            
            # TABLEAU DE DONNÉES
            st.markdown("### 📋 Données détaillées")
            
            df_display = df_grouped.copy()
            df_display['VAR_MONTANT'] = df_display['MONTANT'].diff()
            df_display['VAR_MONTANT_%'] = (df_display['MONTANT'].pct_change() * 100).fillna(0)
            df_display['VAR_CONSO'] = df_display['CONSO'].diff()
            df_display['VAR_CONSO_%'] = (df_display['CONSO'].pct_change() * 100).fillna(0)
            
            df_display = df_display[['DATE_DISPLAY', 'MONTANT', 'VAR_MONTANT', 'VAR_MONTANT_%', 
                                   'CONSO', 'VAR_CONSO', 'VAR_CONSO_%']]
            df_display.columns = ['Période', 'Montant', 'Δ Montant', 'Δ % Montant', 
                                'Conso', 'Δ Conso', 'Δ % Conso']
            
            st.dataframe(
                df_display.style.format({
                    'Montant': '{:,.0f} FCFA',
                    'Δ Montant': '{:+,.0f} FCFA',
                    'Δ % Montant': '{:+.1f}%',
                    'Conso': '{:,.0f} kWh',
                    'Δ Conso': '{:+,.0f} kWh',
                    'Δ % Conso': '{:+.1f}%'
                }),
                use_container_width=True,
                height=300
            )

elif page == "⚙️ Génération Fichiers":
    st.markdown("## ⚙️ Génération des Fichiers Comptables")
    st.markdown("---")
    
    # Tabs pour BT et HT
    tab_bt, tab_ht = st.tabs(["🔌 Basse Tension (BT)", "⚡ Haute Tension (HT)"])
    
    with tab_bt:
        st.markdown("### 📋 Génération Fichier CIE BT")
        st.markdown("*Sélectionnez une période pour générer le fichier avec les montants à jour*")
        
        # Charger le template BT
        df_template_bt = load_template_bt()
        
        if df_template_bt is None:
            st.error("❌ Template BT introuvable : FACTURAT_ELECTRICITE_BT.xlsx")
        else:
            # Filtre de date
            df_central = st.session_state.df_central
            
            if 'DATE' in df_central.columns:
                periodes_disponibles = sorted(df_central['DATE'].dropna().unique().tolist(), reverse=True)
                
                if len(periodes_disponibles) > 0:
                    col_f1, col_f2 = st.columns(2)
                    
                    with col_f1:
                        periode_selectionnee = st.selectbox(
                            "📅 Sélectionner la période",
                            periodes_disponibles,
                            help="Choisissez la période pour générer le fichier"
                        )
                    
                    with col_f2:
                        if periode_selectionnee:
                            st.info(f"📊 Période sélectionnée : **{periode_selectionnee}**")
                    
                    st.markdown("---")
                    
                    # Filtrer la base centrale par la période sélectionnée
                    df_central_filtre = df_central[df_central['DATE'] == periode_selectionnee]

                    # DÉBOGAGE AJOUTÉ
                    st.info(f"""
                    🔍 **Informations de débogage** :
                    - Période sélectionnée : `{periode_selectionnee}`
                    - Nombre de lignes dans base centrale filtrée : {len(df_central_filtre)}
                    - Exemples de dates dans base centrale : {df_central_filtre['DATE'].head(3).tolist() if len(df_central_filtre) > 0 else 'Aucune'}
                    """)
                    
                    if len(df_central_filtre) > 0:
                        with st.expander("👁️ Voir la base centrale filtrée (débogage)"):
                            st.dataframe(df_central_filtre[['IDENTIFIANT', 'SITES', 'MONTANT', 'DATE']].head(10), use_container_width=True)

                    if len(df_central_filtre) == 0:
                        st.warning(f"⚠️ Aucune donnée pour la période {periode_selectionnee}")
                    else:
                        # Créer une copie du template pour mise à jour
                        df_export_bt = df_template_bt.copy()
                        
                        nb_maj = 0
                        
                        # Pour chaque ligne du template
                        for idx, row_template in df_export_bt.iterrows():
                            identifiant_template = normaliser_identifiant(row_template['IDENTIFIANT'])
                            
                            # Chercher dans la base centrale filtrée
                            ligne_centrale = df_central_filtre[df_central_filtre['IDENTIFIANT'] == identifiant_template]
                            
                            if not ligne_centrale.empty:
                                montant_trouve = ligne_centrale.iloc[0]['MONTANT']
                                
                                # Vérifier que le montant n'est pas vide
                                if pd.notna(montant_trouve) and montant_trouve != '':
                                    # Mettre à jour le MONTANT
                                    df_export_bt.at[idx, 'MONTANT'] = montant_trouve
                                    
                                    # Mettre à jour le LIBELLE COMPLEMENTAIRE
                                    site = ligne_centrale.iloc[0]['SITES'] if 'SITES' in ligne_centrale.columns else ''
                                    df_export_bt.at[idx, 'LIBELLE COMPLEMENTAIRE'] = f"CIE BT {periode_selectionnee} {site}"
                                    
                                    nb_maj += 1
                        
                        # Afficher les statistiques
                        col_stat1, col_stat2, col_stat3 = st.columns(3)
                        with col_stat1:
                            st.metric("📝 Lignes template", len(df_export_bt))
                        with col_stat2:
                            st.metric("✅ Lignes mises à jour", nb_maj)
                        with col_stat3:
                            total_bt = pd.to_numeric(df_export_bt['MONTANT'], errors='coerce').fillna(0).sum()
                            st.metric("💰 Total", f"{total_bt:,.0f} FCFA")
                        
                        st.markdown("---")
                        
                        # Aperçu
                        with st.expander("👁️ Aperçu du fichier à générer"):
                            st.dataframe(df_export_bt, use_container_width=True)
                        
                        st.markdown("---")
                        
                        # Boutons d'action
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            if st.button("💾 Sauvegarder template", use_container_width=True, key="save_bt_gen"):
                                df_export_bt.to_excel(FICHIER_TEMPLATE_BT, index=False)
                                st.success("✅ Template BT sauvegardé !")
                                st.rerun()
                        
                        with col2:
                            # Supprimer IDENTIFIANT pour l'export
                            df_final_bt = df_export_bt.drop(columns=['IDENTIFIANT']) if 'IDENTIFIANT' in df_export_bt.columns else df_export_bt
                            
                            # Export avec signatures - style BT bleu
                            excel_data_bt = export_factures_cie(df_final_bt, type_tension="BT")
                            
                            st.download_button(
                                "📥 Télécharger Excel BT avec signatures",
                                data=excel_data_bt,
                                file_name=f"FACTURAT_ELECTRICITE_BT_{periode_selectionnee}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="dl_bt_gen"
                            )
                        
                        with col3:
                            if st.button("🔄 Actualiser", use_container_width=True, key="refresh_bt_gen"):
                                st.rerun()
                else:
                    st.warning("⚠️ Aucune période disponible dans la base centrale")
            else:
                st.warning("⚠️ La colonne DATE n'existe pas dans la base centrale")
    
    with tab_ht:
        st.markdown("### 📋 Génération Fichier CIE HT")
        st.markdown("*Sélectionnez une période pour générer le fichier avec les montants à jour*")
        
        # Charger le template HT
        df_template_ht = load_template_ht()
        
        if df_template_ht is None:
            st.error("❌ Template HT introuvable : FACTURAT_ELECTRICITE_HT.xlsx")
        else:
            # Filtre de date
            df_central = st.session_state.df_central
            
            if 'DATE' in df_central.columns:
                periodes_disponibles = sorted(df_central['DATE'].dropna().unique().tolist(), reverse=True)
                
                if len(periodes_disponibles) > 0:
                    col_f1, col_f2 = st.columns(2)
                    
                    with col_f1:
                        periode_selectionnee = st.selectbox(
                            "📅 Sélectionner la période",
                            periodes_disponibles,
                            help="Choisissez la période pour générer le fichier",
                            key="periode_ht"
                        )
                    
                    with col_f2:
                        if periode_selectionnee:
                            st.info(f"📊 Période sélectionnée : **{periode_selectionnee}**")
                    
                    st.markdown("---")
                    
                    # Filtrer la base centrale par la période sélectionnée
                    df_central_filtre = df_central[df_central['DATE'] == periode_selectionnee]
                    
                    # DÉBOGAGE AJOUTÉ
                    st.info(f"""
                    🔍 **Informations de débogage** :
                    - Période sélectionnée : `{periode_selectionnee}`
                    - Nombre de lignes dans base centrale filtrée : {len(df_central_filtre)}
                    - Exemples de dates dans base centrale : {df_central_filtre['DATE'].head(3).tolist() if len(df_central_filtre) > 0 else 'Aucune'}
                    """)
                    
                    if len(df_central_filtre) > 0:
                        with st.expander("👁️ Voir la base centrale filtrée (débogage)"):
                            st.dataframe(df_central_filtre[['IDENTIFIANT', 'SITES', 'MONTANT', 'DATE']].head(10), use_container_width=True)
                    
                    if len(df_central_filtre) == 0:
                        st.warning(f"⚠️ Aucune donnée pour la période {periode_selectionnee}")
                    else:
                        # Créer une copie du template pour mise à jour
                        df_export_ht = df_template_ht.copy()
                        
                        nb_maj = 0
                        
                        # Pour chaque ligne du template
                        for idx, row_template in df_export_ht.iterrows():
                            identifiant_template = normaliser_identifiant(row_template['IDENTIFIANT'])
                            
                            # Chercher dans la base centrale filtrée
                            ligne_centrale = df_central_filtre[df_central_filtre['IDENTIFIANT'] == identifiant_template]
                            
                            if not ligne_centrale.empty:
                                montant_trouve = ligne_centrale.iloc[0]['MONTANT']
                                
                                # Vérifier que le montant n'est pas vide
                                if pd.notna(montant_trouve) and montant_trouve != '':
                                    # Mettre à jour le MONTANT
                                    df_export_ht.at[idx, 'MONTANT'] = montant_trouve
                                    
                                    # Mettre à jour le LIBELLE COMPLEMENTAIRE
                                    site = ligne_centrale.iloc[0]['SITES'] if 'SITES' in ligne_centrale.columns else ''
                                    df_export_ht.at[idx, 'LIBELLE COMPLEMENTAIRE'] = f"CIE HT {periode_selectionnee} {site}"
                                    
                                    nb_maj += 1
                        
                        # Afficher les statistiques
                        col_stat1, col_stat2, col_stat3 = st.columns(3)
                        with col_stat1:
                            st.metric("📝 Lignes template", len(df_export_ht))
                        with col_stat2:
                            st.metric("✅ Lignes mises à jour", nb_maj)
                        with col_stat3:
                            total_ht = pd.to_numeric(df_export_ht['MONTANT'], errors='coerce').fillna(0).sum()
                            st.metric("💰 Total", f"{total_ht:,.0f} FCFA")
                        
                        st.markdown("---")
                        
                        # Aperçu
                        with st.expander("👁️ Aperçu du fichier à générer"):
                            st.dataframe(df_export_ht, use_container_width=True)
                        
                        st.markdown("---")
                        
                        # Boutons d'action
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            if st.button("💾 Sauvegarder template", use_container_width=True, key="save_ht_gen"):
                                df_export_ht.to_excel(FICHIER_TEMPLATE_HT, index=False)
                                st.success("✅ Template HT sauvegardé !")
                                st.rerun()
                        
                        with col2:
                            # Supprimer IDENTIFIANT pour l'export
                            df_final_ht = df_export_ht.drop(columns=['IDENTIFIANT']) if 'IDENTIFIANT' in df_export_ht.columns else df_export_ht
                            
                            # Export avec signatures - style HT orange
                            excel_data_ht = export_factures_cie(df_final_ht, type_tension="HT")
                            
                            st.download_button(
                                "📥 Télécharger Excel HT avec signatures",
                                data=excel_data_ht,
                                file_name=f"FACTURAT_ELECTRICITE_HT_{periode_selectionnee}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="dl_ht_gen"
                            )
                        
                        with col3:
                            if st.button("🔄 Actualiser", use_container_width=True, key="refresh_ht_gen"):
                                st.rerun()
                else:
                    st.warning("⚠️ Aucune période disponible dans la base centrale")
            else:
                st.warning("⚠️ La colonne DATE n'existe pas dans la base centrale")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; padding: 2rem; color: #666;'>
    <p><strong>Système de traitement des factures CIE/SODECI</strong> - Version 2.0</p>
</div>
""", unsafe_allow_html=True)
